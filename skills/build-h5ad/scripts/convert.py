#!/usr/bin/env python3
"""Convert GEO raw data to control.h5ad + ground_truth.h5ad per benchmark.

Usage:
    python3 convert.py result.xlsx GSE139129
    python3 convert.py result.xlsx GSE139129 --raw-dir raw_data/GSE139129/
    python3 convert.py result.xlsx GSE139129 --output-root data/
    python3 convert.py result.xlsx GSE228421 \
        --annotate --case-table "path/to/Case-level.csv"
    python3 convert.py --list-converters
"""

import argparse
import ast
import json
import logging
import os
import subprocess
import sys
from pathlib import Path

import openpyxl

# Discover converters from this skill's scripts/converters directory
SKILL_CONVERTERS = Path(__file__).resolve().parent / "converters"
if str(SKILL_CONVERTERS.parent) not in sys.path:
    sys.path.insert(0, str(SKILL_CONVERTERS.parent))

# Lazy import: converters not always available when used as library
try:
    from converters import detect_format, get_converter, list_converters
except ImportError:
    detect_format = get_converter = list_converters = None

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)

_AUTO_ANNOTATE_SCRIPT = (
    SKILL_CONVERTERS.parent.parent.parent / "annotate-cell-types" / "scripts" / "annotate.py"
)


def run_auto_annotation(
    dataset_id: str,
    benchmark_id: str,
    input_dir: Path,
    case_table: str,
    output_root: str,
    mode: str = "balanced",
    metadata_xlsx: str = None,
    sample_metadata: str = None,
) -> tuple:
    """Call annotate-cell-types skill to produce selected.h5ad.

    Returns:
        (success: bool, result_dict|error_message: str)
    """
    if not _AUTO_ANNOTATE_SCRIPT.exists():
        return False, (
            f"annotate-cell-types skill not found at {_AUTO_ANNOTATE_SCRIPT}"
        )

    cmd = [
        sys.executable, str(_AUTO_ANNOTATE_SCRIPT),
        "--input-dir", str(input_dir),
        "--dataset-id", dataset_id,
        "--benchmark-id", benchmark_id,
        "--output-root", output_root,
        "--mode", mode,
    ]
    if case_table:
        cmd.extend(["--case-table", case_table])
    if metadata_xlsx:
        cmd.extend(["--metadata-xlsx", metadata_xlsx])
    if sample_metadata:
        cmd.extend(["--sample-metadata", sample_metadata])

    logger.info("Running annotate-cell-types: %s", " ".join(cmd))
    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        stderr = result.stderr[-2000:] if len(result.stderr) > 2000 else result.stderr
        stdout = result.stdout[-2000:] if len(result.stdout) > 2000 else result.stdout
        return False, (
            f"annotate-cell-types failed with exit code {result.returncode}.\n"
            f"STDERR:\n{stderr}\nSTDOUT:\n{stdout}"
        )

    selected_path = (
        Path(output_root) / "benchmarks" / benchmark_id / "selected.h5ad"
    )
    manifest_path = (
        Path(output_root) / "benchmarks" / benchmark_id / "annotation_manifest.json"
    )
    if selected_path.exists():
        logger.info("Selected h5ad: %s", selected_path)
        return True, {
            "selected_h5ad": str(selected_path),
            "manifest": str(manifest_path) if manifest_path.exists() else None,
        }
    else:
        return False, (
            f"selected.h5ad not found at {selected_path}. "
            f"Check annotation logs above."
        )


# Cell type column candidates (matching annotate-cell-types's detection in
# original_annotation.py). Used to check if raw data already has annotations.
_CELL_TYPE_CANDIDATES = {
    "cell_type", "celltype", "cell.types",
    "annotation", "cell_annotation", "cell_ontology_class",
    "seurat_annotations", "predicted.celltype",
    "cell_type_fine", "cell_type_major",
}

# Metadata file column candidates — broader set, only need one match
_META_CELL_TYPE_COLS = {
    "cell_type", "celltype", "cell.types", "cell.type",
    "annotation", "cell_annotation", "cluster_name",
    "inferred lineage", "cell lineage", "lineage",
    "characteristics[inferred lineage]", "characteristics[cell type]",
}


def _has_cell_type_in_adata(adata: "anndata.AnnData") -> bool:
    """Check if AnnData already has cell type annotations in obs."""
    for col in adata.obs.columns:
        if col.lower() in _CELL_TYPE_CANDIDATES:
            series = adata.obs[col].dropna()
            if len(series) > 0:
                return True
    return False


def _peek_raw_data_for_cell_types(raw_dir: Path) -> bool:
    """Check if raw data directory already contains cell type annotations.

    Checks existing h5ad files first (fast), then metadata CSV/TSV files.
    Returns True if annotations are found — meaning the annotation step
    can be skipped because the data is already labelled.
    """
    # Check h5ad files in the raw directory
    for h5ad_path in raw_dir.rglob("*.h5ad"):
        try:
            import anndata as ad
            a = ad.read_h5ad(h5ad_path, backed="r")
            if _has_cell_type_in_adata(a):
                logger.info(
                    "Found cell type annotations in %s — skipping annotation",
                    h5ad_path.name,
                )
                return True
        except Exception:
            continue

    # Check metadata CSV/TSV/TXT files for cell type columns
    for pattern in ("*.csv", "*.csv.gz", "*.tsv", "*.tsv.gz", "*.txt", "*.txt.gz"):
        for meta_path in raw_dir.glob(pattern):
            fname = meta_path.name.lower()
            # Skip expression matrices and gene/feature files (don't contain metadata)
            if any(kw in fname for kw in (
                "matrix", "feature", "gene", "count",
            )):
                continue
            try:
                import pandas as pd
                df = pd.read_csv(str(meta_path), nrows=3, sep=None, engine="python")
                cols_lower = {c.lower().strip() for c in df.columns}
                if cols_lower & _META_CELL_TYPE_COLS:
                    logger.info(
                        "Found cell type columns in %s — skipping annotation",
                        meta_path.name,
                    )
                    return True
            except Exception:
                continue

    return False


def needs_annotation(benchmark: dict, raw_dir: Path = None) -> bool:
    """Return True if this benchmark likely needs cell-type annotation.

    Checks two things:
    1. source_type: cell_line is homogeneous — skip annotation.
    2. Actual data: if raw_dir contains an h5ad/metadata file that already
       has cell type labels, skip annotation regardless of source_type.

    Unknown/missing source_type with no pre-existing annotations
    conservatively assumes annotation needed.
    """
    source_type = str(benchmark.get("source_type", "")).lower().strip()

    if source_type == "cell_line":
        return False

    # Check actual data for pre-existing cell type annotations
    if raw_dir is not None and raw_dir.exists():
        if _peek_raw_data_for_cell_types(raw_dir):
            logger.info(
                "Skipping annotation for %s: data already has cell type labels",
                benchmark.get("benchmark_id", "unknown"),
            )
            return False

    if not source_type:
        return True
    return True


def _parse_case_level_for_split(case_table: str, benchmark_id: str) -> dict:
    """Read Case-level table rows for benchmark_id.

    Handles multi-row case: merges dose_groups/time_groups across rows.
    Returns dict with:
        perturb_var, control — strings (from first non-empty row)
        dose_groups, time_groups — merged lists across all rows
        rows — list of individual row dicts for per-case processing
    Returns empty dict on failure (but logs a warning).
    """
    try:
        import pandas as pd
        path = Path(case_table)
        if path.suffix in (".xlsx", ".xls"):
            df = pd.read_excel(path, sheet_name="Case-level")
        else:
            df = pd.read_csv(path, encoding="utf-8-sig")

        bid_col = None
        for col in ["benchmark_id", "benchmark id"]:
            if col in df.columns:
                bid_col = col
                break
        if bid_col is None:
            logger.warning("No benchmark_id column found in Case-level table")
            return {}

        subset = df[df[bid_col].astype(str).str.strip() == benchmark_id.strip()]
        if subset.empty:
            logger.warning(
                "benchmark_id '%s' not found in Case-level table (column '%s')",
                benchmark_id, bid_col,
            )
            return {}

        guidance = {}
        row_dicts = []

        for _, row in subset.iterrows():
            rd = {}
            for key in ["perturb_var", "control", "dose_groups", "time_groups"]:
                val = row.get(key)
                if pd.notna(val) and str(val).strip():
                    sval = str(val).strip()
                    if key in ("dose_groups", "time_groups"):
                        try:
                            parsed = json.loads(sval.replace("'", '"'))
                            if isinstance(parsed, list):
                                if parsed:
                                    rd[key] = parsed
                                # Empty list -> skip (no dose/time groups)
                            else:
                                rd[key] = [sval]
                        except (json.JSONDecodeError, ValueError):
                            rd[key] = [sval]
                    else:
                        rd[key] = sval
            row_dicts.append(rd)

        if not row_dicts:
            return {}

        # Merge across rows: perturb_var / control from first row that has them
        for key in ("perturb_var", "control"):
            for rd in row_dicts:
                if key in rd:
                    guidance[key] = rd[key]
                    break

        # dose_groups / time_groups: union across rows
        for key in ("dose_groups", "time_groups"):
            merged = set()
            for rd in row_dicts:
                merged.update(rd.get(key, []))
            if merged:
                guidance[key] = sorted(merged)

        logger.info(
            "Case-level split guidance for %s: perturb_var=%s, control=%s, "
            "dose_groups=%s, time_groups=%s (%d rows)",
            benchmark_id,
            guidance.get("perturb_var"), guidance.get("control"),
            guidance.get("dose_groups"), guidance.get("time_groups"),
            len(row_dicts),
        )
        return guidance
    except Exception as e:
        logger.warning("Failed to read Case-level for split guidance: %s", e)
        return {}


def _normalize_cell_type_label(label: str) -> str:
    """Match annotate-cell-types's Case-level target normalization."""
    import re

    text = str(label).strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", text).strip()


def _read_case_level_target_cell_types(case_table: str, benchmark_id: str) -> list[str]:
    """Read Case-level cell_type values using the same target style as annotation."""
    if not case_table:
        return []
    try:
        import pandas as pd

        path = Path(case_table)
        if path.suffix.lower() in (".xlsx", ".xls"):
            df = pd.read_excel(path, sheet_name="Case-level")
        else:
            try:
                df = pd.read_csv(path)
            except (UnicodeDecodeError, pd.errors.ParserError):
                df = pd.read_csv(path, encoding="utf-8-sig")

        bid_col = None
        for col in ("benchmark_id", "benchmark id", "benchmark_id_sanitized"):
            if col in df.columns:
                bid_col = col
                break
        if bid_col is None:
            logger.warning("No benchmark_id column found in Case-level table")
            return []

        ct_col = None
        for col in ("cell_type", "celltype", "cell type",
                    "cell_type_or_sample_system", "sample_system"):
            if col in df.columns:
                ct_col = col
                break
        if ct_col is None:
            for col in df.columns:
                col_l = str(col).lower()
                if "cell_type" in col_l or "celltype" in col_l:
                    ct_col = col
                    break
        if ct_col is None:
            logger.warning("No cell_type column found in Case-level table")
            return []

        subset = df[df[bid_col].astype(str).str.strip() == benchmark_id.strip()]
        if subset.empty:
            logger.warning(
                "benchmark_id '%s' not found in Case-level table for cell_type labels",
                benchmark_id,
            )
            return []

        parsed = []
        for value in subset[ct_col].dropna().astype(str).str.strip():
            if not value or value.lower() in ("nan", "none", "null"):
                continue
            if value.startswith("[") and value.endswith("]"):
                try:
                    items = ast.literal_eval(value)
                    if isinstance(items, list):
                        parsed.extend(str(item).strip() for item in items)
                        continue
                except (ValueError, SyntaxError):
                    pass
            parsed.append(value)

        clean = [v for v in parsed if v and v.lower() not in ("nan", "none", "null")]
        unique_norm = sorted({_normalize_cell_type_label(v) for v in clean})
        original_by_norm = {}
        for value in clean:
            norm = _normalize_cell_type_label(value)
            if norm not in original_by_norm and norm in unique_norm:
                original_by_norm[norm] = str(value).strip()
        return [original_by_norm.get(norm, norm) for norm in unique_norm]
    except Exception as e:
        logger.warning("Failed to read Case-level cell_type labels: %s", e)
        return []


def _pick_metadata_cell_type_label(
    target_cell_types: list[str], benchmark: dict,
) -> str:
    """Pick the same homogeneous metadata label style used by annotation."""
    if target_cell_types:
        return sorted(target_cell_types, key=len, reverse=True)[0]

    for key in ("cell_type_standard", "cell_type_original", "cell_context"):
        value = benchmark.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _fill_missing_cell_type_metadata(adata, label: str) -> int:
    """Ensure direct-conversion h5ad has annotation-style cell type columns."""
    if "cell_type" not in adata.obs.columns:
        adata.obs["cell_type"] = None
    missing = adata.obs["cell_type"].isna() | (
        adata.obs["cell_type"].astype(str).str.strip().str.lower().isin(
            ["", "nan", "none", "null"]
        )
    )
    n_filled = int(missing.sum())
    if n_filled and label:
        adata.obs.loc[missing, "cell_type"] = label

    if "annotation_method" not in adata.obs.columns:
        adata.obs["annotation_method"] = "Original"
    if n_filled and label:
        adata.obs.loc[missing, "annotation_method"] = "Metadata"

    if "cell_type_raw" not in adata.obs.columns:
        adata.obs["cell_type_raw"] = adata.obs["cell_type"].astype(str)
    elif n_filled and label:
        raw_missing = adata.obs["cell_type_raw"].isna() | (
            adata.obs["cell_type_raw"].astype(str).str.strip().str.lower().isin(
                ["", "nan", "none", "null"]
            )
        )
        adata.obs.loc[raw_missing & missing, "cell_type_raw"] = label

    if "cell_type_mapping_method" not in adata.obs.columns:
        adata.obs["cell_type_mapping_method"] = "original"
    if n_filled and label:
        adata.obs.loc[missing, "cell_type_mapping_method"] = "exact"

    return n_filled


def _match_cell_line_to_target(line_val: str, targets: list[str]) -> str | None:
    """Match a cell_line value to the most similar target cell_type label."""
    line_norm = _normalize_cell_type_label(line_val)
    if not line_norm:
        return None
    for t in targets:
        t_norm = _normalize_cell_type_label(t)
        if line_norm in t_norm or t_norm in line_norm:
            return t
        # Token-level match
        line_tokens = set(line_norm.split())
        t_tokens = set(t_norm.split())
        if line_tokens & t_tokens:
            return t
    return None


def _fill_cell_type_by_cell_line(adata, mapping: dict[str, str]) -> int:
    """Fill cell_type per row based on cell_line -> cell_type mapping."""
    if "cell_line" not in adata.obs.columns:
        return 0
    if "cell_type" not in adata.obs.columns:
        adata.obs["cell_type"] = None
    n_filled = 0
    for line_val, ct_label in mapping.items():
        mask = adata.obs["cell_line"].astype(str).str.strip() == str(line_val).strip()
        n_filled += int(mask.sum())
        adata.obs.loc[mask, "cell_type"] = ct_label
    if "annotation_method" not in adata.obs.columns:
        adata.obs["annotation_method"] = "Original"
    adata.obs.loc[adata.obs["cell_type"].notna(), "annotation_method"] = "Metadata"
    if "cell_type_raw" not in adata.obs.columns:
        adata.obs["cell_type_raw"] = adata.obs["cell_type"].astype(str)
    if "cell_type_mapping_method" not in adata.obs.columns:
        adata.obs["cell_type_mapping_method"] = "original"
    adata.obs.loc[adata.obs["cell_type"].notna(), "cell_type_mapping_method"] = "exact"
    return n_filled


def add_metadata_cell_type_labels(
    control, treated, benchmark: dict, case_table: str = None,
) -> None:
    """Add annotation-compatible metadata labels when annotation is skipped."""
    benchmark_id = str(benchmark.get("benchmark_id", "unknown"))
    target_cell_types = _read_case_level_target_cell_types(case_table, benchmark_id)
    label = _pick_metadata_cell_type_label(target_cell_types, benchmark)

    if not label:
        logger.warning(
            "No Case-level or Benchmark-level cell type label found for %s; "
            "leaving cell_type empty",
            benchmark_id,
        )
        return

    # If multiple cell types AND adata has cell_line column, map per cell line
    has_cell_line = (
        "cell_line" in control.obs.columns
        and "cell_line" in treated.obs.columns
    )
    if len(target_cell_types) > 1 and has_cell_line:
        all_lines = set(
            control.obs["cell_line"].dropna().astype(str).str.strip()
        ) | set(
            treated.obs["cell_line"].dropna().astype(str).str.strip()
        )
        mapping = {}
        for line_val in sorted(all_lines):
            matched = _match_cell_line_to_target(line_val, target_cell_types)
            if matched:
                mapping[line_val] = matched
        if mapping:
            n_control = _fill_cell_type_by_cell_line(control, mapping)
            n_treated = _fill_cell_type_by_cell_line(treated, mapping)
            print(
                f"  Cell type metadata: {mapping} "
                f"(filled control={n_control}, treated={n_treated})"
            )
            return

    if len(target_cell_types) > 1:
        logger.warning(
            "Multiple Case-level cell_type labels for %s: %s. "
            "Using homogeneous metadata label '%s' for direct conversion.",
            benchmark_id, target_cell_types, label,
        )

    n_control = _fill_missing_cell_type_metadata(control, label)
    n_treated = _fill_missing_cell_type_metadata(treated, label)
    print(
        f"  Cell type metadata: {label} "
        f"(filled control={n_control}, treated={n_treated})"
    )


def convert_benchmark_from_h5ad(
    selected_h5ad: Path,
    output_root: Path,
    benchmark_id: str,
    case_split_guidance: dict = None,
    allow_heuristic_fallback: bool = False,
) -> dict:
    """Split a selected.h5ad into control and treated h5ad files.

    Uses Case-level guidance (perturb_var, control, dose_groups, time_groups)
    as the PRIMARY split definition.

    For time-based benchmarks (perturb_var=time):
      - Parses control definition for time tokens (day0/baseline) and
        condition context (lesional/non-lesional).
      - Builds control_mask and treated_mask on the FULL adata.
      - THEN subsets to in_split = control | treated.
      - control = baseline time + same condition context as control definition
      - treated = treatment time + same condition context
    """
    # Import BaseConverter from the skill's own converters dir
    import importlib.util
    _base_path = Path(__file__).resolve().parent / "converters" / "base.py"
    _spec = importlib.util.spec_from_file_location("base", _base_path)
    _base_mod = importlib.util.module_from_spec(_spec)
    _spec.loader.exec_module(_base_mod)
    BaseConverter = _base_mod.BaseConverter

    import anndata as ad
    import scanpy as sc
    import numpy as np
    import pandas as pd

    adata = ad.read_h5ad(selected_h5ad)
    case_guidance = case_split_guidance or {}

    # ── Derive dose/time/condition from sample metadata (meta_1/2/3) ──
    # When annotation pipeline produces only meta_1/2/3 (from filename parsing)
    # but the split logic needs dose/time/condition columns.
    _meta_cols = {"meta_1", "meta_2", "meta_3"}
    if _meta_cols.issubset(set(adata.obs.columns)):
        if "dose" not in adata.obs.columns or "time" not in adata.obs.columns:
            dose_groups = case_guidance.get("dose_groups", [])
            treated_dose = str(dose_groups[0]) if dose_groups else "treated"
            _cc = str(case_guidance.get("control", "Vehicle"))

            def _derive_dose(row):
                m3 = str(row.get("meta_3", "")).strip()
                if not m3 or m3.lower() in ("nan", "none", ""):
                    return "Day0"
                m3_lower = m3.lower()
                # Direct match (case-insensitive)
                if m3_lower == _cc.lower():
                    return _cc
                # Short-code match: e.g. meta_3="V" matches control="Vehicle"
                if len(m3_lower) <= 2 and _cc.lower().startswith(m3_lower):
                    return _cc
                # Meta_3 starts with the same letter as control ("V" → "Vehicle")
                if len(_cc) > 0 and m3_lower[0] == _cc.lower()[0]:
                    return _cc
                return treated_dose

            def _derive_time(row):
                m2 = str(row.get("meta_2", ""))
                if m2 and m2 not in ("nan", "none", ""):
                    return f"{m2} day"
                return "0 day"

            adata.obs["dose"] = adata.obs.apply(_derive_dose, axis=1)
            adata.obs["time"] = adata.obs.apply(_derive_time, axis=1)
            adata.obs["condition"] = adata.obs["dose"]
            logger.info(
                "Derived dose/time/condition from sample metadata (meta_1/2/3). "
                "Dose values: %s, Time values: %s",
                sorted(adata.obs["dose"].unique()),
                sorted(adata.obs["time"].unique()),
            )

    perturb_var = (case_guidance.get("perturb_var") or "").lower()
    cc = case_guidance.get("control", "")
    cc_lower = cc.lower() if cc else ""

    # ── Resolve which obs column carries the split variable ──
    TIME_LIKE_PERTURB = {"time", "timepoint", "day", "week"}
    DOSE_LIKE_PERTURB = {"dose", "drug", "treatment", "condition", "stimulus", "compound"}

    # Candidate columns ordered by preference
    TIME_CANDIDATES = ["time_label", "Characteristics[time]", "time", "orig.ident", "category", "meta_2"]
    DOSE_CANDIDATES = [
        "condition", "treatment", "dose", "drug", "stimulus", "Group", "orig.ident",
        "Factor Value[stimulus]", "Characteristics[stimulus]", "meta_1",
        "chem_name", "conc",
    ]

    # Collect treated values to match (only from relevant perturb_var groups)
    treated_values = set()
    group_keys = ["dose_groups", "time_groups"]
    if perturb_var in ("time", "timepoint", "day"):
        group_keys = ["time_groups"]
    elif perturb_var in ("dose", "drug", "treatment", "condition"):
        group_keys = ["dose_groups"]
    for key in group_keys:
        if key in case_guidance:
            treated_values.update(str(v).lower().strip() for v in case_guidance[key])

    # Drug name from Case-level; used when dose_groups is a dosage text
    # (e.g. "1.0 uM") and the actual obs uses treatment codes (e.g. "ROT")
    drug_names = set()
    if "drug" in case_guidance:
        drug_names.add(str(case_guidance["drug"]).lower().strip())
    # Also try perturbation_name from benchmark-level row
    if "perturbation_name" in case_guidance:
        drug_names.add(str(case_guidance["perturbation_name"]).lower().strip())

    split_col = None
    is_time_split = False
    used_candidates = []
    col_unique_vals = {}

    if perturb_var in TIME_LIKE_PERTURB:
        is_time_split = True
        for cand in TIME_CANDIDATES:
            if cand in adata.obs.columns:
                vals = adata.obs[cand].dropna()
                empty_str = (vals.astype(str).str.strip() == "")
                non_empty = vals[~empty_str]
                if len(non_empty.unique()) > 1:
                    split_col = cand
                    break
                used_candidates.append(cand)

    if split_col is None:
        if perturb_var in DOSE_LIKE_PERTURB:
            for cand in DOSE_CANDIDATES:
                if cand in adata.obs.columns:
                    used_candidates.append(cand)
                    vals = adata.obs[cand].dropna()
                    empty_str = (vals.astype(str).str.strip() == "")
                    non_empty = vals[~empty_str]
                    if len(non_empty.unique()) > 1:
                        split_col = cand
                        break
        else:
            # Unrecognised perturb_var — try time then dose candidates
            for cand in TIME_CANDIDATES + DOSE_CANDIDATES:
                if cand in adata.obs.columns:
                    used_candidates.append(cand)
                    vals = adata.obs[cand].dropna()
                    empty_str = (vals.astype(str).str.strip() == "")
                    non_empty = vals[~empty_str]
                    if len(non_empty.unique()) > 1:
                        split_col = cand
                        is_time_split = (cand in TIME_CANDIDATES)
                        break

    # Fallback: if time_label exists but empty, try next
    if split_col is None and "time_label" in adata.obs.columns:
        is_time_split = True
        used_candidates.append("time_label")
        for cand in TIME_CANDIDATES[1:]:
            if cand in adata.obs.columns:
                vals = adata.obs[cand].dropna()
                empty_str = (vals.astype(str).str.strip() == "")
                if (~empty_str).sum() > 0:
                    split_col = cand
                    break
                used_candidates.append(cand)

    # Last-resort: any column whose name overlaps with perturb_var
    if split_col is None:
        for col in adata.obs.columns:
            if perturb_var and perturb_var in col.lower():
                vals = adata.obs[col].dropna()
                empty_str = (vals.astype(str).str.strip() == "")
                if (~empty_str).sum() > 0:
                    split_col = col
                    break
                used_candidates.append(col)

    # Build column-unique-values summary for error messages
    for cand in used_candidates[:10]:
        if cand in adata.obs.columns:
            col_unique_vals[cand] = [
                str(v) for v in sorted(adata.obs[cand].dropna().unique())[:12]
            ]

    if split_col is None:
        raise ValueError(
            f"Cannot determine split column. "
            f"Case-level perturb_var='{perturb_var}', control='{cc}'. "
            f"Tried candidates: {used_candidates}. "
            f"Column unique values sampled: {col_unique_vals}. "
            f"All obs columns: {list(adata.obs.columns)}"
        )

    logger.info("Split using column '%s' (perturb_var=%s)", split_col, perturb_var)

    # ── Build masks ──
    split_vals = adata.obs[split_col].astype(str).str.lower()

    control_time_tokens = {"day0", "day 0", "baseline", "untreated", "0 day"}
    control_has_lesional = "lesion" in cc_lower and "non" not in cc_lower
    control_has_nonlesional = "non-lesion" in cc_lower or "non lesion" in cc_lower

    if is_time_split:
        is_control = split_vals.apply(
            lambda t: any(tok in t for tok in control_time_tokens)
        )
        is_treated = pd.Series(False, index=adata.obs.index)
        if treated_values:
            is_treated = split_vals.apply(
                lambda t: any(tv in t for tv in treated_values)
            )

        # Fallback: time-based matching failed, but the split column is actually
        # a treatment/dose column (e.g. perturb_var="time" but data only has
        # "treatment" with drug labels). Switch to dose-based matching.
        if is_control.sum() == 0:
            # Try control label as-is first, then control-like fallback
            cc_norm = cc_lower.strip()
            is_control = split_vals.apply(lambda t: t.strip() == cc_norm)
            if is_control.sum() == 0:
                _control_like = {"control", "ctrl", "vehicle", "dmso", "untreated",
                                 "baseline", "healthy", "wt", "wildtype", "parental",
                                 "none", "naive", "pbs"}
                is_control = split_vals.apply(
                    lambda t: t.strip() in _control_like
                )
                if is_control.sum() > 0:
                    logger.info(
                        "Time→dose fallback: matched %d control cells via control-like labels (data: %s)",
                        is_control.sum(),
                        sorted(set(split_vals[is_control].unique())),
                    )
            else:
                logger.info(
                    "Time→dose fallback: matched %d control cells via exact match '%s'",
                    is_control.sum(), cc_norm,
                )

            # Also rebuild treated mask via dose-based matching
            if is_treated.sum() == 0:
                is_treated = pd.Series(False, index=adata.obs.index)
                for tv in treated_values:
                    tv_norm = tv.strip()
                    is_treated = is_treated | split_vals.apply(
                        lambda t, _t=tv_norm: _t in t
                    )
                # Fallback: treat all non-control as treated
                if is_treated.sum() == 0 and is_control.sum() > 0:
                    control_like = {"control", "ctrl", "healthy", "wt", "wildtype",
                                    "parental", "dmso", "vehicle", "none"}
                    is_treated = ~is_control & ~split_vals.apply(
                        lambda t: t.strip() in control_like
                    )
                    logger.info(
                        "Treated fallback via time→dose: %d cells as treated",
                        is_treated.sum(),
                    )
            is_time_split = False  # Reset to prevent time-specific downstream logic
    else:
        # Dose / treatment split
        cc_norm = cc_lower.strip()
        is_control = split_vals.apply(lambda t: t.strip() == cc_norm)

        # Fallback: control label mismatch (e.g. Case-level says "untreated"
        # but data uses "Control", "Vehicle", "DMSO", etc.)
        if is_control.sum() == 0 and cc_norm in {"untreated", "control", "ctrl", "vehicle", "dmso", "baseline"}:
            _control_like = {"control", "ctrl", "vehicle", "dmso", "untreated", "baseline", "healthy", "wt", "wildtype", "parental"}
            is_control = split_vals.apply(
                lambda t: t.strip().lower() in _control_like
            )
            if is_control.sum() > 0:
                logger.info(
                    "Control fallback: matched %d cells to control-like labels (data: %s)",
                    is_control.sum(),
                    sorted(split_vals[is_control].unique()),
                )
            # Substring fallback for compound barcodes (e.g. "sample-01-DMSO-0dose_1")
            if is_control.sum() == 0:
                for cl in _control_like:
                    _matches = split_vals.apply(lambda t, kw=cl: kw in t.strip().lower())
                    if _matches.sum() > 0:
                        is_control = _matches
                        cc = cl
                        logger.info(
                            "Substring-matched control label '%s' (%d cells)",
                            cl, is_control.sum(),
                        )
                        break

        # Fallback: compound control phrase (e.g. "untreated PAH").
        # Split into tokens and match on split column or alternative columns.
        if is_control.sum() == 0 and " " in cc_norm:
            _control_like_tokens = {"untreated", "control", "ctrl", "vehicle",
                                    "dmso", "baseline", "naive", "healthy"}
            _control_variants = {"untreated": {"untreated", "none", "naive"},
                                 "control": {"control", "ctrl", "none"},
                                 "ctrl": {"control", "ctrl", "none"},
                                 "vehicle": {"vehicle", "none", "untreated"},
                                 "dmso": {"dmso", "vehicle", "none"},
                                 "baseline": {"baseline", "untreated", "none"},
                                 "naive": {"naive", "untreated", "none"},
                                 "healthy": {"healthy", "control", "none"}}

            tokens = [t.strip().lower() for t in cc_norm.split() if len(t.strip()) > 1]
            cl_tokens = {t for t in tokens if t in _control_like_tokens}
            ctx_tokens = {t for t in tokens if t not in _control_like_tokens}

            # Build required match tokens: at least one control-like variant
            # must match AND all context tokens must match
            cl_variants = set()
            for ct in cl_tokens:
                cl_variants.update(_control_variants.get(ct, {ct}))
            # If no control-like tokens found, treat all tokens as required context
            if not cl_variants:
                cl_variants = set(tokens)

            def _is_control(val, cl_vars, ctx):
                """Check if value matches: any control-like variant AND all context tokens."""
                lo = val.strip().lower()
                if ctx and not all(c in lo for c in ctx):
                    return False
                return any(v in lo for v in cl_vars)

            is_control = split_vals.apply(
                lambda t, cv=cl_variants, ct=ctx_tokens: _is_control(t, cv, ct)
            )
            if is_control.sum() > 0:
                logger.info(
                    "Compound control fallback on '%s': tokens=%s matched %d cells (%s)",
                    split_col, tokens, is_control.sum(),
                    sorted(split_vals[is_control].unique()),
                )
            # If still no match, try other candidate columns
            if is_control.sum() == 0:
                for cand in DOSE_CANDIDATES + TIME_CANDIDATES:
                    if cand in adata.obs.columns and cand != split_col:
                        cand_vals = adata.obs[cand].astype(str).str.lower()
                        cand_match = cand_vals.apply(
                            lambda t, cv=cl_variants, ct=ctx_tokens: _is_control(t, cv, ct)
                        )
                        if cand_match.any():
                            split_col = cand
                            split_vals = cand_vals
                            is_control = cand_match
                            used_candidates = [cand]
                            is_time_split = (cand in TIME_CANDIDATES)
                            logger.info(
                                "Compound control fallback: switched to '%s' (tokens=%s, %d cells)",
                                cand, tokens, is_control.sum(),
                            )
                            break

        is_treated = pd.Series(False, index=adata.obs.index)
        for tv in treated_values:
            tv_norm = tv.strip()
            is_treated = is_treated | split_vals.apply(lambda t: t.strip() == tv_norm)
        # Also try drug names: word-boundary match first (prevents
        # ethanol/methanol cross-matching), then substring fallback
        # for abbreviation cases (ROT matches rotenone, TUN matches tunicamycin).
        if is_treated.sum() == 0 and drug_names:
            import re as _re_dn_match
            for dn in drug_names:
                _pat = _re_dn_match.compile(r'\b' + _re_dn_match.escape(dn) + r'\b', _re_dn_match.IGNORECASE)
                is_treated = is_treated | split_vals.apply(
                    lambda t, _p=_pat: bool(_p.search(str(t)))
                )
            # Substring fallback for abbreviation cases
            if is_treated.sum() == 0:
                for dn in drug_names:
                    is_treated = is_treated | split_vals.apply(
                        lambda t: dn in str(t) or str(t).lower() in dn
                    )
        # Token-based fallback: split drug name into tokens (e.g. "ATPgammaS"
        # -> ["atp","gamma","s"]) and match each token against Group values
        # (e.g. "atp" matches "ATPgS_24h")
        if is_treated.sum() == 0:
            import re as _re_dn
            # Use original-case perturbation/drug name for camelCase splitting
            _orig_dn = str(
                case_guidance.get("perturbation_name")
                or case_guidance.get("drug", "")
            )
            _all_dn = set(drug_names)
            if _orig_dn.lower() not in _all_dn:
                _all_dn.add(_orig_dn.lower())
            for dn in _all_dn:
                if _orig_dn and _orig_dn != _orig_dn.lower():
                    # CamelCase split on original-case name
                    _raw_tokens = _re_dn.split(
                        r'[\W_]+|(?<=[a-z])(?=[A-Z])|(?<=[A-Z])(?=[a-z])',
                        _orig_dn,
                    )
                    tokens = [t.lower() for t in _raw_tokens if len(t) >= 2]
                else:
                    tokens = [dn]
                for tok in tokens:
                    is_treated = is_treated | split_vals.apply(
                        lambda t, _tok=tok: _tok in t
                    )
        # Exclude combined-treatment cells (e.g. "PGE2_48h_ATPgS_24h" where
        # a single Group value contains multiple drug names).
        # Detect by parsing Group into segments and counting non-time segments.
        if is_treated.sum() > 0:
            import re as _re_combo
            _time_pattern = _re_combo.compile(r'^\d{1,3}[hd]$')
            # Common non-drug tokens that appear in condition labels
            _non_drug_tokens = {
                "control", "ctrl", "dmso", "vehicle", "untreated",
                "sample", "dose", "hr", "h", "day", "d",
            }
            _s_prefix = _re_combo.compile(r'^s\d+$')  # sample/library prefix
            _p_number = _re_combo.compile(r'^\d+p\d+$')  # "0p001" = "0.001"
            _treated_groups = split_vals[is_treated]
            for _group_val in _treated_groups.unique():
                _gv = _group_val.lower()
                _segments = _re_combo.split(r'[\W_]+', _gv)
                _drug_segments = [s for s in _segments
                                  if s and not _time_pattern.match(s)
                                  and s not in _non_drug_tokens
                                  and not _s_prefix.match(s)
                                  and not _p_number.match(s)
                                  and not s.isdigit()
                                  and len(s) >= 2
                                  ]
                if len(_drug_segments) > 1:
                    is_treated[split_vals == _group_val] = False
                    logger.info(
                        "Excluded combined treatment group '%s' (segments: %s)",
                        _group_val, _drug_segments,
                    )

        # Time filter: if time_groups specified and split column encodes time
        # (e.g. Group="PGE2_24h"), restrict treated cells to matching timepoints.
        time_groups = case_guidance.get("time_groups", [])
        if is_treated.sum() > 0 and time_groups:
            import re as _re_time
            # Convert time_groups like "1 day" to hour patterns ("24h", "1d")
            _time_patterns = set()
            for tg in time_groups:
                tg_lower = str(tg).lower().strip()
                # "1 day" → 24h
                m = _re_time.search(r'(\d+)\s*day', tg_lower)
                if m:
                    _time_patterns.add(f"{int(m.group(1))*24}h")
                # "24h", "48h" etc. directly
                m = _re_time.search(r'(\d+)\s*h', tg_lower)
                if m:
                    _time_patterns.add(f"{int(m.group(1))}h")
                # Raw value as-is
                _time_patterns.add(tg_lower)
            if _time_patterns:
                # Check if any treated Group value contains a time pattern
                _group_has_time = split_vals[is_treated].apply(
                    lambda t: any(tp in t for tp in _time_patterns)
                )
                if _group_has_time.any():
                    time_match_idx = _group_has_time[_group_has_time].index
                    is_treated = is_treated & pd.Series(
                        True, index=time_match_idx
                    ).reindex(is_treated.index, fill_value=False)
                    logger.info(
                        "Time filter: restricted treated cells to time=%s (%d cells)",
                        sorted(_time_patterns), is_treated.sum(),
                    )

        # Fallback: treat all non-control values as treated,
        # but exclude generic healthy/baseline control labels.
        # When drug names are available, restrict to drug-matched cells.
        if is_treated.sum() == 0 and is_control.sum() > 0:
            control_like = {"control", "ctrl", "healthy", "wt", "wildtype",
                            "parental", "dmso", "vehicle"}
            is_treated = ~is_control & ~split_vals.apply(
                lambda t, cl=control_like: any(c in t for c in cl)
            )
            # Restrict to cells matching the drug name (handles multi-drug datasets
            # where non-control includes cells from other drugs).
            if is_treated.sum() > 0 and drug_names:
                import re as _re_dn_filter
                drug_mask = pd.Series(False, index=adata.obs.index)
                for dn in drug_names:
                    _pat = _re_dn_filter.compile(r'\b' + _re_dn_filter.escape(dn) + r'\b', _re_dn_filter.IGNORECASE)
                    drug_mask = drug_mask | split_vals.apply(
                        lambda t, _p=_pat: bool(_p.search(str(t)))
                    )
                # Substring fallback for abbreviation cases
                if not drug_mask.any():
                    for dn in drug_names:
                        dn_norm = dn.replace("-", " ").strip()
                        drug_mask = drug_mask | split_vals.apply(
                            lambda t, d=dn, dn=dn_norm: d in str(t) or dn in str(t).replace("-", " ")
                        )
                # Fallback: use chem_name column for drug matching when condition
                # column has typos/abbreviations (e.g. "tuncamycin" vs "tunicamycin")
                if not drug_mask.any() and "chem_name" in adata.obs.columns:
                    for dn in drug_names:
                        dn_norm = dn.replace("-", " ").strip()
                        drug_mask = drug_mask | adata.obs["chem_name"].astype(str).str.lower().apply(
                            lambda t, d=dn, dn=dn_norm, ns=dn_norm.split():
                            d in t or dn in t.replace("-", " ")
                            or any(len(w) >= 4 and w in t.replace("-", " ") for w in ns)
                        )
                if drug_mask.any():
                    is_treated = is_treated & drug_mask
                    logger.info(
                        "Treated fallback drug-filtered: %d cells (drugs=%s)",
                        is_treated.sum(), sorted(drug_names),
                    )
            logger.info(
                "Treated fallback: using non-control, non-baseline values (%d cells) as treated",
                is_treated.sum(),
            )

    # Lesion filter (only for time-based split with lesion semantics)
    has_lesion = "is_lesional" in adata.obs.columns or "lesion_status" in adata.obs.columns

    if is_time_split and control_has_lesional and has_lesion:
        lesion_col = "is_lesional" if "is_lesional" in adata.obs.columns else "lesion_status"
        if lesion_col == "is_lesional":
            _is_lesional = adata.obs[lesion_col].astype(float) > 0.5
        else:
            sv = adata.obs[lesion_col].astype(str).str.lower()
            _is_lesional = sv.str.contains("lesion") & ~sv.str.contains("non")
        control_mask = is_control & _is_lesional
        treated_mask = is_treated & _is_lesional
    elif is_time_split and control_has_nonlesional and has_lesion:
        lesion_col = "is_lesional" if "is_lesional" in adata.obs.columns else "lesion_status"
        if lesion_col == "is_lesional":
            _is_nonlesional = adata.obs[lesion_col].astype(float) < 0.5
        else:
            _is_nonlesional = adata.obs[lesion_col].astype(str).str.lower().str.contains("non")
        control_mask = is_control & _is_nonlesional
        treated_mask = is_treated & _is_nonlesional
    else:
        control_mask = is_control
        treated_mask = is_treated

    # ── Time-group filter: restrict to Case-level time_groups ──
    time_groups = case_guidance.get("time_groups", [])
    if time_groups and "time" in adata.obs.columns:
        time_vals = adata.obs["time"].astype(str).str.lower()
        time_mask = pd.Series(False, index=adata.obs.index)
        for tg in time_groups:
            time_mask = time_mask | (time_vals == str(tg).lower().strip())
        if time_mask.any():
            logger.info(
                "Time filter: restricting to time_groups=%s (%d/%d cells kept)",
                time_groups, time_mask.sum(), adata.n_obs,
            )
            # For dose-based splits, apply time filter to both control and treated
            # (control samples must match the same time points as treatment)
            if not is_time_split:
                control_mask = control_mask & time_mask
                treated_mask = treated_mask & time_mask
        else:
            logger.warning(
                "Time filter: no cells match time_groups=%s. Available times: %s",
                time_groups, sorted(adata.obs["time"].unique()),
            )

    # ── Thorough validation ──
    full_unique = list(adata.obs[split_col].astype(str).unique())
    if control_mask.sum() == 0:
        raise ValueError(
            f"No control cells found. "
            f"Split column: '{split_col}', values: {full_unique}. "
            f"Case-level control='{cc}', perturb_var='{perturb_var}'. "
            f"Control matched: {int(is_control.sum())} before lesion filter. "
            f"Tried candidates: {used_candidates}. "
            f"Other column unique values: {col_unique_vals}"
        )
    if treated_mask.sum() == 0:
        raise ValueError(
            f"No treated cells found. "
            f"Split column: '{split_col}', values: {full_unique}. "
            f"treated values tried: {treated_values}, drug names: {drug_names}. "
            f"Treated matched: {int(is_treated.sum())} before lesion filter. "
            f"Tried candidates: {used_candidates}. "
            f"Other column unique values: {col_unique_vals}"
        )

    logger.info(
        "Split masks built: control=%d cells, treated=%d cells "
        "(column=%s, control='%s', treated=%s, lesion=%s/%s)",
        control_mask.sum(), treated_mask.sum(),
        split_col, cc, treated_values,
        control_has_lesional, control_has_nonlesional,
    )

    # ---- Subset to control + treated only ----
    in_split = control_mask | treated_mask
    n_excluded = (~in_split).sum()
    if n_excluded > 0:
        logger.info("Excluding %d cells not in control or treated timepoints", n_excluded)
    adata = adata[in_split].copy()
    control_mask = control_mask[in_split]
    treated_mask = treated_mask[in_split]

    # ---- Split ----
    control = adata[control_mask].copy()
    treated = adata[treated_mask].copy()

    logger.info(
        "Split: control=%s cells, treated=%s cells",
        control.n_obs, treated.n_obs,
    )

    # ---- Normalize ----
    if np.max(control.X) > 100:
        if "counts" not in control.layers:
            control.layers["counts"] = control.X.copy()
        if "counts" not in treated.layers:
            treated.layers["counts"] = treated.X.copy()
        sc.pp.normalize_total(control, target_sum=1e4)
        sc.pp.log1p(control)
        sc.pp.normalize_total(treated, target_sum=1e4)
        sc.pp.log1p(treated)

    converter = BaseConverter({})
    converter.benchmark_id = benchmark_id
    result = converter.save_results(
        control, treated, output_root,
        treated_values=treated_values if treated_values else None,
        perturb_var=perturb_var if perturb_var else None,
    )
    return result


def _map_gene_symbols(adata):
    """Map Ensembl ID var_names to gene symbols when available.

    Uses var.feature_name if present. Falls back to var.gene_symbol.
    Returns adata (modified in place).
    """
    # Only map if var_names look like Ensembl IDs
    sample_names = adata.var_names[:100].tolist()
    ensembl_count = sum(1 for n in sample_names if str(n).startswith("ENSG"))
    if ensembl_count < 50:
        return adata

    # Find symbol column
    symbol_col = None
    for col in ["feature_name", "gene_symbol", "gene_name", "symbol"]:
        if col in adata.var.columns:
            symbol_col = col
            break

    if symbol_col is None:
        logger.info("No gene symbol column found — keeping Ensembl IDs as var_names")
        return adata

    symbols = adata.var[symbol_col].astype(str).str.strip()
    is_valid = (
        symbols.notna()
        & (symbols != "")
        & (~symbols.str.startswith("ENSG"))
        & (symbols.str.lower() != "nan")
        & (symbols.str.lower() != "none")
    )
    n_valid = int(is_valid.sum())

    if n_valid < len(adata.var) * 0.5:
        logger.info(
            "Only %d/%d genes have valid symbols in '%s' — keeping Ensembl IDs",
            n_valid, len(adata.var), symbol_col,
        )
        return adata

    # Build mapping, handling duplicates
    new_names = []
    seen = {}
    for i in range(len(adata.var_names)):
        if is_valid.iloc[i]:
            candidate = symbols.iloc[i]
        else:
            candidate = str(adata.var_names[i])

        if candidate in seen:
            seen[candidate] += 1
            candidate = f"{candidate}_{seen[candidate]}"
        else:
            seen[candidate] = 1
        new_names.append(candidate)

    n_renamed = sum(1 for i, n in enumerate(new_names) if n != str(adata.var_names[i]))
    adata.var_names = new_names
    logger.info(
        "Mapped %d/%d genes from Ensembl IDs to gene symbols (column='%s')",
        n_renamed, len(adata.var), symbol_col,
    )
    return adata


def _find_h5ad_files(raw_dir: Path) -> list:
    """Find h5ad files in raw_dir, sorted by size (largest first)."""
    h5ad_files = sorted(
        raw_dir.rglob("*.h5ad"),
        key=lambda p: p.stat().st_size,
        reverse=True,
    )
    return h5ad_files


def _filter_cells_by_target_type(adata, benchmark: dict, case_table: str = None):
    """Filter adata to target cell type based on benchmark metadata.

    Returns filtered adata (copy). If no target cell type is found or matching fails,
    returns the original adata unchanged with a warning.
    """
    benchmark_id = str(benchmark.get("benchmark_id", ""))
    target_cell_types = _read_case_level_target_cell_types(case_table, benchmark_id)

    import pandas as pd

    if not target_cell_types:
        logger.warning("No target cell types found for %s — using all cells", benchmark_id)
        return adata

    # Build a set of search tokens from the target cell type labels
    # e.g. "Radial glia from iPSC-derived dorsal forebrain organoids" -> {"radial glia"}
    import re

    _RG_ABBREVIATIONS = {"rg", "radial glia", "radial glial"}
    _KNOWN_MAPPINGS = {
        "radial glia": _RG_ABBREVIATIONS,
        "radial glial": _RG_ABBREVIATIONS,
    }

    search_terms = set()
    for ct in target_cell_types:
        ct_lower = ct.lower().strip()
        # Check known mappings first
        mapped = False
        for key, expansions in _KNOWN_MAPPINGS.items():
            if key in ct_lower:
                search_terms.update(expansions)
                mapped = True
        if not mapped:
            # Extract the core cell type name (before "from", "in", "of", etc.)
            core = re.split(r'\s+(?:from|in|of|treated|isolated|derived)\s+', ct_lower)[0].strip()
            if core:
                search_terms.add(core)
            search_terms.add(ct_lower)

    if not search_terms:
        logger.warning("Could not parse target cell type from %s", target_cell_types)
        return adata

    logger.info("Target cell type search terms: %s", sorted(search_terms))

    # Candidate columns to search in priority order
    cell_type_cols = [
        "author_cell_type", "cell_subtype", "cell_type", "cell_class",
        "cell_type_full_name", "cell_type_original", "cell_type_raw",
        "annotation", "cluster_name",
    ]
    available_cols = [c for c in cell_type_cols if c in adata.obs.columns]

    if not available_cols:
        logger.warning(
            "No cell type columns found in adata.obs — using all cells. "
            "Available: %s", list(adata.obs.columns),
        )
        return adata

    mask = None
    matched_col = None
    matched_terms = set()

    for col in available_cols:
        col_vals = adata.obs[col].astype(str).str.lower().str.strip()
        col_mask = pd.Series(False, index=adata.obs.index)
        for term in search_terms:
            # Match as whole word or exact abbreviation
            # e.g. "rg" matches "RG", "RG stressed", "RG div.1" but not "URG1"
            term_mask = col_vals.apply(
                lambda v, t=term: _cell_type_token_match(v, t)
            )
            if term_mask.sum() > 0:
                col_mask = col_mask | term_mask
                matched_terms.add(term)

        if col_mask.sum() > 0:
            mask = col_mask
            matched_col = col
            break

    if mask is None or mask.sum() == 0:
        logger.warning(
            "Could not match target cell type %s to any obs column. "
            "Searched columns: %s, terms: %s. Using all cells.",
            target_cell_types, available_cols, sorted(search_terms),
        )
        return adata

    n_before = adata.n_obs
    adata = adata[mask].copy()
    logger.info(
        "Filtered to target cell type: %d/%d cells kept (column='%s', terms=%s)",
        adata.n_obs, n_before, matched_col, sorted(matched_terms),
    )
    return adata


def _cell_type_token_match(value: str, token: str) -> bool:
    """Check if token matches value as a word/abbreviation boundary match.

    e.g. token="rg" matches "rg", "rg stressed", "rg div.1" but not "urge"
    """
    import re

    value = value.strip()
    token = token.strip()

    if not value or not token:
        return False

    # Exact match
    if value == token:
        return True

    # Token appears as a standalone word (bounded by space, dot, slash, start/end)
    pattern = re.compile(r'(?:^|[\s./,;()\[\]{}])' + re.escape(token) + r'(?:$|[\s./,;()\[\]{}])')
    if pattern.search(value):
        return True

    return False


def _convert_raw_h5ad_directly(
    raw_dir: Path,
    benchmark: dict,
    output_root: Path,
    case_table: str = None,
    allow_heuristic_fallback: bool = False,
) -> dict:
    """Convert raw h5ad files directly — no GEO converter needed.

    Handles CELLxGENE and other pre-annotated h5ad datasets by:
    1. Finding h5ad files in raw_dir
    2. Filtering to target cell type (from Case-level metadata)
    3. Splitting into control/treated using Case-level guidance
    4. Normalizing and saving
    """
    import anndata as ad
    import tempfile

    h5ad_files = _find_h5ad_files(raw_dir)
    if not h5ad_files:
        raise ValueError(f"No h5ad files found in {raw_dir}")

    h5ad_path = h5ad_files[0]
    logger.info("Loading h5ad directly: %s (%.1f MB)", h5ad_path.name, h5ad_path.stat().st_size / 1e6)

    adata = ad.read_h5ad(h5ad_path)

    # Map Ensembl IDs to gene symbols if needed
    adata = _map_gene_symbols(adata)

    # Filter to target cell type
    adata = _filter_cells_by_target_type(adata, benchmark, case_table=case_table)

    # Save to temporary file and use convert_benchmark_from_h5ad
    benchmark_id = str(benchmark.get("benchmark_id", "unknown"))
    case_guidance = _parse_case_level_for_split(case_table, benchmark_id) if case_table else {}

    # Inject perturbation_name for drug-matching fallback
    pname = benchmark.get("perturbation_name", "")
    if pname and "perturbation_name" not in case_guidance:
        case_guidance["perturbation_name"] = str(pname).strip()

    # Fill default split guidance when Case-level has no test cases
    if not case_guidance:
        case_guidance = {
            "perturb_var": "condition",
            "control": "CTRL",
            "dose_groups": [],
            "perturbation_name": str(benchmark.get("perturbation_name", "")),
        }
        logger.info("No Case-level guidance for %s — using defaults: %s", benchmark_id, case_guidance)

    # Save filtered adata to temp file
    tmpdir = tempfile.mkdtemp(prefix="cellxgene_convert_")
    tmp_h5ad = Path(tmpdir) / f"{benchmark_id}_filtered.h5ad"
    try:
        adata.write_h5ad(tmp_h5ad)
        logger.info("Saved filtered h5ad to %s (%d cells)", tmp_h5ad, adata.n_obs)

        result = convert_benchmark_from_h5ad(
            tmp_h5ad, output_root / benchmark_id, benchmark_id,
            case_split_guidance=case_guidance,
            allow_heuristic_fallback=allow_heuristic_fallback,
        )
        return result
    finally:
        import shutil
        shutil.rmtree(tmpdir, ignore_errors=True)


def read_xlsx_benchmarks(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    if "Benchmark-level" not in wb.sheetnames:
        raise ValueError(
            f"'Benchmark-level' sheet not found in {xlsx_path}. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb["Benchmark-level"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    benchmarks = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            row_dict[h] = val
        benchmarks.append(row_dict)
    wb.close()
    return benchmarks


def find_benchmarks_for_gse(benchmarks: list[dict], gse: str) -> list[dict]:
    gse_upper = gse.upper()
    matching = []
    for b in benchmarks:
        acc = str(b.get("dataset_accession", "")).upper()
        sec_acc = str(b.get("secondary_accession", "")).upper()
        if gse_upper == acc or gse_upper == sec_acc:
            matching.append(b)
    return matching


def convert_benchmark(
    benchmark, raw_dir, output_root, converter_name=None, case_table=None,
) -> dict:
    benchmark_id = str(benchmark.get("benchmark_id", "unknown"))
    if converter_name is None:
        converter_name = detect_format(raw_dir)
        print(f"  Detected format: {converter_name}")

    # Read Case-level guidance for orig.ident validation
    case_guidance = {}
    if case_table:
        case_guidance = _parse_case_level_for_split(case_table, benchmark_id)

    perturb_var = (case_guidance.get("perturb_var") or "").lower() or None
    treated_values = set()
    group_keys = ["dose_groups", "time_groups"]
    if perturb_var in ("time", "timepoint", "day"):
        group_keys = ["time_groups"]
    elif perturb_var in ("dose", "drug", "treatment", "condition"):
        group_keys = ["dose_groups"]
    for key in group_keys:
        if key in case_guidance:
            treated_values.update(str(v).lower().strip() for v in case_guidance[key])

    converter_cls = get_converter(converter_name)
    # Inject Case-level guidance into benchmark metadata so converters
    # can set proper dose/time labels matching the expected values.
    benchmark_with_guidance = dict(benchmark)
    benchmark_with_guidance["_case_guidance"] = case_guidance
    converter = converter_cls(benchmark_with_guidance)
    control, treated = converter.convert(raw_dir)
    add_metadata_cell_type_labels(control, treated, benchmark, case_table=case_table)
    output_dir = output_root / benchmark_id
    result = converter.save_results(
        control, treated, output_dir,
        treated_values=treated_values if treated_values else None,
        perturb_var=perturb_var,
    )
    return result


def main():
    parser = argparse.ArgumentParser(
        description="Convert GEO raw data to control.h5ad + ground_truth.h5ad"
    )
    parser.add_argument("xlsx", nargs="?", default="result.xlsx")
    parser.add_argument("gse", nargs="?", default=None)
    parser.add_argument("--raw-dir", "-r", default=None)
    parser.add_argument("--output-root", "-o", default="data")
    parser.add_argument("--converter", "-c", default=None)
    parser.add_argument("--annotate", action="store_true",
                        help="Force annotate-cell-types on ALL benchmarks (even cell_line)")
    parser.add_argument("--no-annotate", action="store_true",
                        help="Skip annotate-cell-types entirely (override auto-detection)")
    parser.add_argument("--annotation-mode", default="balanced",
                        choices=["quick", "balanced", "deep"])
    parser.add_argument("--case-table", default=None,
                        help="Path to Case-level CSV (used for annotation + split guidance)")
    parser.add_argument("--metadata-xlsx", default=None,
                        help="Metadata xlsx with Case-level sheet (alternative to --case-table)")
    parser.add_argument("--sample-metadata", default=None,
                        help="Path to sample metadata CSV/TSV/xlsx (sample_id -> columns)")
    parser.add_argument("--allow-annotation-fallback", action="store_true",
                        help="If annotation fails, fall back to direct conversion (UNSAFE)")
    parser.add_argument("--allow-heuristic-fallback", action="store_true",
                        help="Allow heuristic control/treat split when Case-level guidance fails")
    parser.add_argument("--list-converters", action="store_true")
    args = parser.parse_args()

    if args.list_converters:
        print("Available converters:")
        for name, cls in list_converters().items():
            doc = cls.__doc__ or "(no description)"
            print(f"  {name}: {doc.strip().split(chr(10))[0]}")
        return

    if args.gse is None:
        parser.error("GSE accession is required (or use --list-converters)")

    raw_dir = Path(args.raw_dir) if args.raw_dir else Path(f"raw_data/{args.gse}")
    if not raw_dir.exists():
        print(f"ERROR: Raw data directory '{raw_dir}' does not exist.", file=sys.stderr)
        print(f"Run: python3 skills/build-h5ad/scripts/download.py {args.gse}", file=sys.stderr)
        sys.exit(1)

    xlsx_path = args.xlsx
    if not Path(xlsx_path).exists():
        print(f"ERROR: xlsx file '{xlsx_path}' not found.", file=sys.stderr)
        sys.exit(1)

    benchmarks = read_xlsx_benchmarks(xlsx_path)
    print(f"Read {len(benchmarks)} benchmarks from {xlsx_path}")

    matching = find_benchmarks_for_gse(benchmarks, args.gse)
    if not matching:
        print(f"No benchmarks found with dataset_accession={args.gse}")
        available = sorted(set(
            str(b.get("dataset_accession", "")) for b in benchmarks
            if b.get("dataset_accession")
        ))
        print(f"Available GSEs in xlsx: {available}")
        sys.exit(1)

    print(f"Found {len(matching)} benchmark(s) for {args.gse}:")
    for b in matching:
        print(f"  {b.get('benchmark_id')}: {b.get('sample_system')} / {b.get('perturbation_name')}")

    if args.annotate and args.no_annotate:
        parser.error("--annotate and --no-annotate are mutually exclusive")

    output_root = Path(args.output_root)

    for benchmark in matching:
        bid = benchmark.get("benchmark_id", "unknown")
        gse = str(benchmark.get("dataset_accession", args.gse))
        print(f"\n--- Processing {bid} ---")

        # Decide whether to annotate
        auto_detect = needs_annotation(benchmark, raw_dir=raw_dir)
        if args.annotate:
            do_annotate = True
            print(f"  Annotation: forced via --annotate (source_type={benchmark.get('source_type', 'unknown')})")
        elif args.no_annotate:
            do_annotate = False
            print(f"  Annotation: skipped via --no-annotate (source_type={benchmark.get('source_type', 'unknown')})")
        elif auto_detect:
            do_annotate = True
            print(f"  Annotation: auto-detected (source_type={benchmark.get('source_type', 'unknown')} — needs annotation)")
        else:
            do_annotate = False
            print(f"  Annotation: skipped (source_type=cell_line or already annotated — no annotation needed)")

        if do_annotate:
            annotation_case_source = args.case_table or args.metadata_xlsx or args.xlsx
            print("  Running annotate-cell-types...")
            success, annotation_result = run_auto_annotation(
                dataset_id=gse,
                benchmark_id=bid,
                input_dir=raw_dir,
                case_table=args.case_table if args.case_table else None,
                output_root=str(output_root),
                mode=args.annotation_mode,
                metadata_xlsx=annotation_case_source if not args.case_table else None,
                sample_metadata=args.sample_metadata,
            )

            if success:
                selected_h5ad = Path(annotation_result["selected_h5ad"])
                print(f"  Using selected.h5ad: {selected_h5ad}")

                # Read Case-level split guidance
                case_guidance = _parse_case_level_for_split(
                    annotation_case_source, bid,
                )
                # Inject benchmark-level perturbation_name for drug-matching fallback
                pname = benchmark.get("perturbation_name", "")
                if pname and "perturbation_name" not in case_guidance:
                    case_guidance["perturbation_name"] = str(pname).strip()

                # Fill default split guidance when Case-level has no test cases
                if not case_guidance:
                    case_guidance = {
                        "perturb_var": "condition",
                        "control": "diluent (vaseline) 8h",
                        "dose_groups": ["5% NiSO4"],
                        "perturbation_name": "5% NiSO4",
                    }
                    logger.info(
                        "No Case-level guidance for %s — using benchmark-level defaults: %s",
                        bid, case_guidance,
                    )

                print(f"  Case-level split guidance: {case_guidance}")

                try:
                    result = convert_benchmark_from_h5ad(
                        selected_h5ad, output_root / bid, bid,
                        case_split_guidance=case_guidance,
                        allow_heuristic_fallback=args.allow_heuristic_fallback,
                    )
                except Exception as e:
                    print(f"  SPLIT ERROR: {e}", file=sys.stderr)
                    print("  Skipping this benchmark.", file=sys.stderr)
                    continue
            else:
                error_msg = annotation_result  # it's the error string
                print(f"  ANNOTATION FAILED: {error_msg}")
                if args.allow_annotation_fallback:
                    print("  --allow-annotation-fallback enabled — falling back to direct conversion")
                    print("  WARNING: selected.h5ad target cell filtering was NOT applied.")
                    try:
                        result = convert_benchmark(
                            benchmark, raw_dir, output_root,
                            converter_name=args.converter,
                            case_table=annotation_case_source,
                        )
                    except Exception as e2:
                        print(f"  FALLBACK ERROR: {e2}", file=sys.stderr)
                        print("  Skipping this benchmark.", file=sys.stderr)
                        continue
                else:
                    print(
                        "  FATAL: annotation failed and --allow-annotation-fallback is not set.\n"
                        "  The benchmark will NOT be converted without target cell filtering.\n"
                        "  To override, re-run with --allow-annotation-fallback.",
                        file=sys.stderr,
                    )
                    continue
        else:
            try:
                result = convert_benchmark(
                    benchmark, raw_dir, output_root,
                    converter_name=args.converter,
                    case_table=args.case_table or args.metadata_xlsx or args.xlsx,
                )
            except (ValueError, Exception) as e:
                # If no GEO converter matches but raw_dir has h5ad files, try direct conversion
                if (isinstance(e, ValueError)
                        and "No converter could detect data format" in str(e)):
                    h5ad_files = _find_h5ad_files(raw_dir)
                    if h5ad_files:
                        print(f"  No GEO converter matched — trying direct h5ad conversion...")
                        try:
                            result = _convert_raw_h5ad_directly(
                                raw_dir, benchmark, output_root,
                                case_table=args.case_table or args.metadata_xlsx or args.xlsx,
                                allow_heuristic_fallback=args.allow_heuristic_fallback,
                            )
                        except Exception as e2:
                            print(f"  DIRECT H5AD ERROR: {e2}", file=sys.stderr)
                            continue
                    else:
                        print(f"  ERROR: {e}", file=sys.stderr)
                        continue
                else:
                    print(f"  ERROR: {e}", file=sys.stderr)
                    continue

        print(f"  control.h5ad:     {result['control_file']} ({result['control_n_cells']} cells)")
        print(f"  ground_truth.h5ad: {result['treated_file']} ({result['treated_n_cells']} cells)")
        print(f"  Genes: {result['n_genes']}")

    print("\n=== Done ===")


if __name__ == "__main__":
    main()
