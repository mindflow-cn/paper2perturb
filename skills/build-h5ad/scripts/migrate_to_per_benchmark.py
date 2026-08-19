#!/usr/bin/env python3
"""Migrate from central test_case.json to per-benchmark test_case.json files.

Reads the central test_cases/test_case.json and writes each benchmark as a
standalone JSON file to data/{benchmark_id}/test_case.json.

For benchmarks that have h5ad files but no entry in the central JSON
(e.g., MB0014BREAST005), generates a skeleton from result.xlsx.

Usage:
    python3 migrate_to_per_benchmark.py
    python3 migrate_to_per_benchmark.py --dry-run
    python3 migrate_to_per_benchmark.py --force
"""

import argparse
import json
import sys
from pathlib import Path

# Fields to include from xlsx benchmark-level when generating skeletons
BENCHMARK_FIELDS = [
    "benchmark_id",
    "sample_system",
    "tissue",
    "source_type",
    "perturbation_type",
    "perturbation_name",
    "default_cell_subset",
    "description",
]

BENCHMARK_OPTIONAL_FIELDS = ["smiles"]

CENTRAL_JSON = Path("test_cases/test_case.json")
DATA_ROOT = Path("data")
XLSX_PATH = Path("result.xlsx")


def load_central_json(path: Path) -> list[dict]:
    """Load benchmark entries from central test_case.json."""
    with open(path) as f:
        data = json.load(f)
    return data.get("benchmarks", [])


def load_xlsx_benchmarks(path: Path) -> list[dict]:
    """Read Benchmark-level sheet from result.xlsx."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True)
    if "Benchmark-level" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Benchmark-level"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if val is None:
                continue
            if isinstance(val, str):
                val = val.strip()
            row_dict[h] = val
        result.append(row_dict)
    return result


def build_skeleton_benchmark(xlsx_row: dict) -> dict:
    """Build a minimal benchmark object from xlsx row (no test cases)."""
    entry = {}
    for field in BENCHMARK_FIELDS:
        val = xlsx_row.get(field)
        entry[field] = str(val).strip() if val is not None else ""

    for field in BENCHMARK_OPTIONAL_FIELDS:
        val = xlsx_row.get(field)
        if val is not None and str(val).strip():
            entry[field] = str(val).strip()

    entry["test_cases"] = []
    return entry


def find_data_dirs(data_root: Path) -> set[str]:
    """Find all directories under data/ that contain h5ad files."""
    dirs = set()
    if not data_root.exists():
        return dirs
    for d in data_root.iterdir():
        if d.is_dir() and not d.name.startswith("."):
            has_h5ad = any(d.glob("*.h5ad"))
            if has_h5ad:
                dirs.add(d.name)
    return dirs


def main():
    parser = argparse.ArgumentParser(
        description="Migrate central test_case.json to per-benchmark files"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview changes without writing files",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Overwrite existing test_case.json files",
    )
    args = parser.parse_args()

    # 1. Load benchmarks from central JSON
    if not CENTRAL_JSON.exists():
        print(f"ERROR: {CENTRAL_JSON} not found", file=sys.stderr)
        sys.exit(1)

    benchmarks = load_central_json(CENTRAL_JSON)
    json_ids = {b["benchmark_id"] for b in benchmarks}
    print(f"Loaded {len(benchmarks)} benchmarks from {CENTRAL_JSON}")

    # 2. Find data dirs with h5ad files
    data_dirs = find_data_dirs(DATA_ROOT)
    print(f"Found {len(data_dirs)} data directories with h5ad files")

    # 3. Load xlsx for skeleton generation
    xlsx_benchmarks = {}
    if XLSX_PATH.exists():
        xlsx_benchmarks = {
            b["benchmark_id"]: b
            for b in load_xlsx_benchmarks(XLSX_PATH)
            if b.get("benchmark_id")
        }
        print(f"Loaded {len(xlsx_benchmarks)} benchmarks from {XLSX_PATH}")

    # 4. Plan: each data dir should get a test_case.json
    created = 0
    skipped = 0
    skeleton_count = 0

    for dir_name in sorted(data_dirs):
        tc_path = DATA_ROOT / dir_name / "test_case.json"

        if tc_path.exists() and not args.force:
            skipped += 1
            print(f"  SKIP {dir_name}/test_case.json (already exists)")
            continue

        entry = None
        if dir_name in json_ids:
            # Find by benchmark_id
            for b in benchmarks:
                if b["benchmark_id"] == dir_name:
                    entry = b
                    break
        elif dir_name in xlsx_benchmarks:
            # Generate skeleton from xlsx
            entry = build_skeleton_benchmark(xlsx_benchmarks[dir_name])
            skeleton_count += 1

        if entry is None:
            print(f"  WARN {dir_name}: no benchmark definition found in JSON or xlsx")
            continue

        if args.dry_run:
            n_cases = len(entry.get("test_cases", []))
            tag = " [SKELETON]" if dir_name not in json_ids else ""
            print(f"  WOULD WRITE {dir_name}/test_case.json ({n_cases} cases){tag}")
            created += 1
        else:
            tc_path.parent.mkdir(parents=True, exist_ok=True)
            tc_path.write_text(json.dumps(entry, indent=2, ensure_ascii=False))
            n_cases = len(entry.get("test_cases", []))
            tag = " [SKELETON]" if dir_name not in json_ids else ""
            print(f"  WROTE {dir_name}/test_case.json ({n_cases} cases){tag}")
            created += 1

    # 5. Report
    print()
    if args.dry_run:
        print(f"DRY RUN: would create {created} files, skip {skipped} existing")
    else:
        print(f"Created {created} files, skipped {skipped} existing")
    if skeleton_count:
        print(f"Includes {skeleton_count} skeleton(s) from xlsx (empty test_cases)")

    # 6. Check for data dirs missing test_case.json definitions
    missing_defs = data_dirs - json_ids - set(xlsx_benchmarks.keys())
    if missing_defs:
        print(f"\nWARNING: {len(missing_defs)} data dir(s) have NO benchmark definition:")
        for d in sorted(missing_defs):
            print(f"  - {d}")
        print("These will be skipped. Add entries to result.xlsx or test_cases/test_case.json.")


if __name__ == "__main__":
    main()
