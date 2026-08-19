#!/usr/bin/env python3
"""Convert result.xlsx benchmark + test case metadata to JSON (test_case.json format).

Usage:
    python3 xlsx2json.py result.xlsx
    python3 xlsx2json.py result.xlsx --output test_cases.json
    python3 xlsx2json.py result.xlsx --gse GSE274905   # filter by dataset_accession
    python3 xlsx2json.py result.xlsx --benchmark-ids MB0001,MB0002
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl


# ---- XLSX column → JSON field mappings ----

# Benchmark-level fields to include in JSON output
BENCHMARK_FIELDS = [
    "benchmark_id",
    "sample_system",
    "tissue",
    "source_type",
    "perturbation_type",
    "perturbation_name",
    "default_cell_subset",
    "description",
]

# Optional benchmark fields (only included when non-empty)
BENCHMARK_OPTIONAL_FIELDS = [
    "smiles",
]

# Case-level fields to include in each test_case
TEST_CASE_FIELDS = [
    "test_id",
    "target_genes",
    "perturb_var",
    "control",
    "dose_groups",
    "time_groups",
    "relation",
    "cell_type",
]

# Case-level fields that need JSON parsing (stored as JSON strings in xlsx)
JSON_PARSE_FIELDS = {"target_genes", "dose_groups", "time_groups"}


def _parse_json_field(value, field_name: str):
    """Parse a JSON string field from xlsx, with error handling."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return [] if field_name in ("dose_groups", "time_groups") else ([] if field_name == "target_genes" else value)
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            print(f"  WARNING: Could not parse {field_name}='{value}' as JSON, keeping as-is",
                  file=sys.stderr)
            return value
    return value


def _cell_value(val):
    """Convert openpyxl cell value to a plain Python value."""
    if val is None:
        return None
    if isinstance(val, str):
        stripped = val.strip()
        return stripped if stripped else None
    # Convert numeric types that are actually integers
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val


def read_xlsx_sheet(xlsx_path: str, sheet_name: str) -> list[dict]:
    """Read a sheet from xlsx and return list of row dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {xlsx_path}. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            row_dict[h] = _cell_value(val)
        result.append(row_dict)

    return result




def build_test_case_json(benchmarks: list[dict], cases: list[dict]) -> dict:
    """Build the test_case.json structure from benchmark and case rows.

    Returns a dict with 'description' and 'benchmarks' keys.
    """
    # Index cases by benchmark_id
    cases_by_benchmark: dict[str, list[dict]] = {}
    for c in cases:
        bid = str(c.get("benchmark_id", ""))
        if not bid:
            continue
        cases_by_benchmark.setdefault(bid, []).append(c)

    benchmark_entries = []
    for b in benchmarks:
        bid = str(b.get("benchmark_id", ""))
        if not bid:
            continue

        entry = {}
        # Required fields
        for field in BENCHMARK_FIELDS:
            val = b.get(field)
            entry[field] = val if val is not None else ""

        # Optional fields (only include if non-empty)
        for field in BENCHMARK_OPTIONAL_FIELDS:
            val = b.get(field)
            if val is not None and str(val).strip():
                entry[field] = str(val).strip()

        # Build test cases for this benchmark
        test_cases = []
        for c in cases_by_benchmark.get(bid, []):
            tc = {}
            for field in TEST_CASE_FIELDS:
                val = c.get(field)
                if field in JSON_PARSE_FIELDS:
                    val = _parse_json_field(val, field)
                elif val is None:
                    val = "" if field in ("control", "perturb_var", "relation") else []
                tc[field] = val
            test_cases.append(tc)

        entry["test_cases"] = test_cases
        benchmark_entries.append(entry)

    return {
        "description": (
            "Dedicated scoring criteria for benchmarks based on real data analysis."
        ),
        "benchmarks": benchmark_entries,
    }


def filter_benchmarks(
    benchmarks: list[dict],
    gse: str = None,
    benchmark_ids: list[str] = None,
) -> list[dict]:
    """Filter benchmarks by GSE accession and/or benchmark IDs."""
    result = benchmarks
    if gse:
        gse_upper = gse.upper()
        result = [
            b for b in result
            if str(b.get("dataset_accession", "")).upper() == gse_upper
            or str(b.get("secondary_accession", "")).upper() == gse_upper
        ]
    if benchmark_ids:
        id_set = set(benchmark_ids)
        result = [b for b in result if str(b.get("benchmark_id", "")) in id_set]
    return result


def main():
    parser = argparse.ArgumentParser(
        description="Convert result.xlsx benchmark metadata to test_case.json format"
    )
    parser.add_argument(
        "xlsx", nargs="?", default="result.xlsx",
        help="Path to result.xlsx (default: result.xlsx)",
    )
    parser.add_argument(
        "--output", "-o", default=None,
        help="Output JSON file path (default: stdout)",
    )
    parser.add_argument(
        "--gse", "-g", default=None,
        help="Filter benchmarks by dataset_accession (e.g., GSE274905)",
    )
    parser.add_argument(
        "--benchmark-ids", "-b", default=None,
        help="Comma-separated benchmark IDs to include",
    )
    parser.add_argument(
        "--per-benchmark-dir", "-d", default=None,
        help="Also write per-benchmark files to {dir}/{benchmark_id}/test_case.json",
    )
    parser.add_argument(
        "--list-benchmarks", action="store_true",
        help="List all benchmark IDs in xlsx and exit",
    )
    args = parser.parse_args()

    xlsx_path = args.xlsx
    if not Path(xlsx_path).exists():
        print(f"ERROR: xlsx file '{xlsx_path}' not found.", file=sys.stderr)
        sys.exit(1)

    # Read both sheets
    benchmarks_raw = read_xlsx_sheet(xlsx_path, "Benchmark-level")
    cases_raw = read_xlsx_sheet(xlsx_path, "Case-level")

    print(f"Read {len(benchmarks_raw)} benchmarks from Benchmark-level sheet",
          file=sys.stderr)
    print(f"Read {len(cases_raw)} test cases from Case-level sheet",
          file=sys.stderr)

    if args.list_benchmarks:
        print("\nBenchmark IDs:")
        for b in benchmarks_raw:
            bid = b.get("benchmark_id", "unknown")
            acc = b.get("dataset_accession", "N/A")
            print(f"  {bid}  (GSE: {acc})")
        return

    # Filter
    benchmark_ids = None
    if args.benchmark_ids:
        benchmark_ids = [x.strip() for x in args.benchmark_ids.split(",")]

    benchmarks = filter_benchmarks(benchmarks_raw, args.gse, benchmark_ids)

    if not benchmarks:
        print("No benchmarks matched the filter criteria.", file=sys.stderr)
        sys.exit(1)

    # Build JSON
    result = build_test_case_json(benchmarks, cases_raw)

    # Summary
    total_cases = sum(len(b["test_cases"]) for b in result["benchmarks"])
    print(
        f"Generated {len(result['benchmarks'])} benchmark(s) "
        f"with {total_cases} test case(s)",
        file=sys.stderr,
    )

    json_str = json.dumps(result, indent=2, ensure_ascii=False)

    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json_str)
        print(f"Wrote {output_path}", file=sys.stderr)
    else:
        print(json_str)

    # Per-benchmark output
    if args.per_benchmark_dir:
        root = Path(args.per_benchmark_dir)
        written = 0
        for entry in result["benchmarks"]:
            bid = entry.get("benchmark_id", "")
            if not bid:
                continue
            bench_dir = root / bid
            bench_dir.mkdir(parents=True, exist_ok=True)
            bench_path = bench_dir / "test_case.json"
            bench_path.write_text(
                json.dumps(entry, indent=2, ensure_ascii=False)
            )
            written += 1
        print(
            f"Wrote {written} per-benchmark test_case.json files to {root}",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
