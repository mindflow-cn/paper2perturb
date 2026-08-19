"""
Deterministic helper functions for benchmark extraction.

Provides CLI subcommands for:
  smiles       — look up SMILES string for a drug name
  pubmed       — fetch publication metadata by PMID
  geo          — fetch GEO series metadata by accession
  gen-benchmark-id — generate next benchmark_id for a PMID
  gen-test-id  — generate next test_id for a benchmark
  write-xlsx   — append benchmark and case rows to result.xlsx (auto-creates if missing)
  assign-ids   — post-hoc assignment of benchmark_ids and test_ids
"""

import argparse
import json
import re
import sys
import time
import xml.etree.ElementTree as ET

import requests

# ---------------------------------------------------------------------------
# Column headers (canonical order)

BENCHMARK_HEADERS = [
    "benchmark_id", "pmid", "paper_doi", "paper_title", "year", "journal",
    "dataset_accession", "secondary_accession", "species",
    "cell_type_original", "cell_type_standard", "cell_type_markers",
    "tissue", "source_type", "cell_context", "disease", "platform",
    "perturbation_type", "perturbation_scope", "perturbation_name",
    "smiles", "description", "control_type", "dose_design", "time_design",
]

CASE_HEADERS = [
    "test_id", "benchmark_id", "pmid", "dataset_accession",
    "source_location", "original_statement", "drug", "cell_type",
    "target_genes", "target_type", "target_id", "perturb_var", "control",
    "dose_groups", "time_groups", "relation", "comparison",
    "gene_set_source", "has_quantitative_support",
    "quantitative_support_type", "quantitative_support_detail",
    "response_timescale", "experiment_design", "is_dose_response",
    "paper_dose_index", "is_time_response", "time_mode",
    "paper_time_index", "notes",
]


# ---------------------------------------------------------------------------
# Init xlsx
# ---------------------------------------------------------------------------

def init_xlsx(result_path: str) -> None:
    """OVERWRITE result.xlsx with fresh Benchmark-level and Case-level sheets + headers only.

    WARNING: All existing data is lost. For appending data, use write_to_xlsx() instead.
    """
    import openpyxl
    import os

    if os.path.exists(result_path):
        print(f"WARNING: Overwriting existing {result_path} — all previous data will be lost.")

    wb = openpyxl.Workbook()
    # First sheet: Benchmark-level
    ws_b = wb.active
    ws_b.title = "Benchmark-level"
    for c, h in enumerate(BENCHMARK_HEADERS, 1):
        ws_b.cell(row=1, column=c, value=h)

    # Second sheet: Case-level
    ws_c = wb.create_sheet("Case-level")
    for c, h in enumerate(CASE_HEADERS, 1):
        ws_c.cell(row=1, column=c, value=h)

    wb.save(result_path)
    print(f"Initialized {result_path} with {len(BENCHMARK_HEADERS)} benchmark columns and {len(CASE_HEADERS)} case columns (all previous data erased)")


# ---------------------------------------------------------------------------
# SMILES lookup
# ---------------------------------------------------------------------------

def fetch_smiles(drug_name: str) -> str:
    """Look up SMILES via PubChem, fallback to ChEMBL."""
    # PubChem
    try:
        resp = requests.get(
            f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{drug_name}/property/CanonicalSMILES/JSON",
            headers={"User-Agent": "PaperFilter/1.0"},
            timeout=10,
        )
        if resp.status_code == 200:
            data = resp.json()
            props = data.get("PropertyTable", {}).get("Properties", [{}])[0]
            # PubChem may return the field as CanonicalSMILES, ConnectivitySMILES, or SMILES
            for key in ("CanonicalSMILES", "ConnectivitySMILES", "SMILES"):
                if key in props:
                    return props[key]
    except Exception:
        pass

    # ChEMBL fallback
    try:
        resp = requests.get(
            "https://www.ebi.ac.uk/chembl/api/data/molecule/search",
            params={"q": drug_name, "format": "json"},
            headers={"User-Agent": "PaperFilter/1.0"},
            timeout=10,
        )
        if resp.status_code == 200:
            data = resp.json()
            molecules = data.get("molecules", [])
            if molecules:
                structs = molecules[0].get("molecule_structures") or {}
                smiles = structs.get("canonical_smiles", "")
                if smiles:
                    return smiles
    except Exception:
        pass

    return ""


# ---------------------------------------------------------------------------
# PubMed metadata
# ---------------------------------------------------------------------------

def fetch_pubmed_metadata(pmid: str) -> dict:
    """Fetch DOI, title, year, journal from PubMed efetch XML API."""
    resp = requests.get(
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi",
        params={"db": "pubmed", "id": pmid, "rettype": "xml"},
        headers={"User-Agent": "PaperFilter/1.0"},
        timeout=15,
    )
    resp.raise_for_status()
    root = ET.fromstring(resp.text)

    result = {"pmid": pmid, "doi": "", "title": "", "year": 0, "journal": ""}

    article = root.find(".//PubmedArticle")
    if article is None:
        return result

    # Title
    title_el = article.find(".//ArticleTitle")
    if title_el is not None and title_el.text:
        result["title"] = title_el.text.strip()

    # DOI
    for aid in article.findall(".//ArticleId"):
        if aid.get("IdType") == "doi" and aid.text:
            result["doi"] = aid.text.strip()
            break

    # Journal
    journal_el = article.find(".//Journal/Title")
    if journal_el is not None and journal_el.text:
        result["journal"] = journal_el.text.strip()

    # Year
    year_el = article.find(".//PubDate/Year")
    if year_el is not None and year_el.text:
        try:
            result["year"] = int(year_el.text)
        except ValueError:
            pass
    if result["year"] == 0:
        medline_el = article.find(".//MedlineDate")
        if medline_el is not None and medline_el.text:
            m = re.search(r"(\d{4})", medline_el.text)
            if m:
                result["year"] = int(m.group(1))

    return result


# ---------------------------------------------------------------------------
# GEO metadata
# ---------------------------------------------------------------------------

def fetch_geo_metadata(gse: str) -> dict:
    """Fetch GEO series metadata via text format API."""
    result = {
        "accession": gse,
        "title": "",
        "platform": "",
        "n_samples": 0,
        "bioproject": "",
        "taxon": "",
        "subseries": [],
    }

    resp = requests.get(
        "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi",
        params={"acc": gse, "targ": "self", "form": "text", "view": "brief"},
        headers={"User-Agent": "PaperFilter/1.0"},
        timeout=15,
    )
    resp.raise_for_status()

    sample_count = 0
    platforms = []
    for line in resp.text.split("\n"):
        line = line.strip()
        if line.startswith("!Series_title"):
            result["title"] = line.split("=", 1)[1].strip()
        elif line.startswith("!Series_platform_id"):
            platforms.append(line.split("=", 1)[1].strip())
        elif line.startswith("!Series_sample_id"):
            sample_count += 1
        elif line.startswith("!Series_platform_organism"):
            result["taxon"] = line.split("=", 1)[1].strip()
        elif "BioProject" in line and "PRJNA" in line:
            m = re.search(r"(PRJNA\d+)", line)
            if m:
                result["bioproject"] = m.group(1)
        elif "SuperSeries of:" in line:
            m = re.search(r"(GSE\d+)", line)
            if m:
                result["subseries"].append(m.group(1))

    result["n_samples"] = sample_count
    result["platform"] = "; ".join(sorted(set(platforms)))

    return result


# ---------------------------------------------------------------------------
# ID generation
# ---------------------------------------------------------------------------

def _read_existing_benchmark_ids(result_path: str) -> list[str]:
    """Read all benchmark_ids from result.xlsx."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(result_path)
    except FileNotFoundError:
        return []

    ws = wb["Benchmark-level"]
    # Find benchmark_id column
    bid_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value and ws.cell(row=1, column=c).value.strip() == "benchmark_id":
            bid_col = c
            break
    if bid_col is None:
        return []

    ids = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=bid_col).value
        if v and str(v).strip():
            ids.append(str(v).strip())
    return ids


def generate_benchmark_id(pmid: str, result_path: str) -> str:
    """Generate next benchmark_id for a given PMID.

    Pattern: <PMID>_<Index:02d>
    - Index starts at 01 per PMID
    """
    existing = _read_existing_benchmark_ids(result_path)
    pmid_str = str(pmid).strip()

    max_idx = 0
    pattern = re.compile(r"^(\d+)_(\d{2})$")

    for bid in existing:
        m = pattern.match(bid)
        if m and m.group(1) == pmid_str:
            idx = int(m.group(2))
            if idx > max_idx:
                max_idx = idx

    return f"{pmid_str}_{max_idx + 1:02d}"


def generate_test_id(benchmark_id: str, result_path: str) -> str:
    """Generate next test_id for a given benchmark.

    Pattern: <benchmark_id>_<Test_Index:02d>
    - Test_Index starts at 01 per benchmark
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(result_path)
    except FileNotFoundError:
        return f"{benchmark_id}_01"

    ws = wb["Case-level"]
    c_test_id_col = _find_header_col(ws, "test_id")
    if c_test_id_col == 0:
        return f"{benchmark_id}_01"

    max_idx = 0
    prefix = f"{benchmark_id}_"
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=c_test_id_col).value
        if v and str(v).strip().startswith(prefix):
            suffix = str(v).strip()[len(prefix):]
            try:
                idx = int(suffix)
                if idx > max_idx:
                    max_idx = idx
            except ValueError:
                pass

    return f"{benchmark_id}_{max_idx + 1:02d}"


# ---------------------------------------------------------------------------
# Write to xlsx
# ---------------------------------------------------------------------------

def write_to_xlsx(result_path: str, benchmark_rows: list[dict], case_rows: list[dict]):
    """Append benchmark and case rows to result.xlsx.

    Auto-creates the file with correct headers if it does not exist.
    """
    import openpyxl
    import os

    if not os.path.exists(result_path):
        init_xlsx(result_path)

    wb = openpyxl.load_workbook(result_path)

    for sheet_name, rows in [("Benchmark-level", benchmark_rows), ("Case-level", case_rows)]:
        if not rows:
            continue
        ws = wb[sheet_name]

        # Build header -> column mapping
        hdr = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v:
                hdr[v.strip()] = c

        # Find last row with data (check column 1)
        last_row = 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is not None:
                last_row = r
        next_row = last_row + 1

        for i, row_data in enumerate(rows):
            target_row = next_row + i
            for key, val in row_data.items():
                col = hdr.get(key.strip())
                if col is None:
                    # Try matching ignoring spaces
                    for h, c in hdr.items():
                        if h.replace(" ", "") == key.replace(" ", ""):
                            col = c
                            break
                if col is None:
                    continue
                # Serialize lists/dicts to JSON strings
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, ensure_ascii=False)
                ws.cell(row=target_row, column=col, value=val)

    wb.save(result_path)
    print(f"Appended {len(benchmark_rows)} benchmarks and {len(case_rows)} cases to {result_path}")


# ---------------------------------------------------------------------------
# Batch ID assignment (for post-hoc ID generation after batch write)
# ---------------------------------------------------------------------------

def _find_header_col(ws, header_name: str) -> int:
    """Find column index (1-based) for a given header name in a worksheet."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v and v.strip() == header_name:
            return c
    return 0


def assign_ids(result_path: str) -> None:
    """Assign benchmark_ids and test_ids to rows that lack them.

    Scans xlsx for benchmark-level rows with empty benchmark_id and case-level
    rows with empty test_id, then computes and fills in proper IDs.  Handles
    batch-write scenarios where IDs were not known at write time.

    benchmark_id pattern: <PMID>_<Index:02d>  (Index starts at 01 per PMID)
    test_id pattern:      <benchmark_id>_<Test_Index:02d>  (starts at 01 per benchmark)
    """
    import openpyxl
    import os

    if not os.path.exists(result_path):
        print(f"Error: {result_path} not found", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(result_path)

    # ---- Benchmark-level ----
    ws_b = wb["Benchmark-level"]
    b_id_col = _find_header_col(ws_b, "benchmark_id")
    b_pmid_col = _find_header_col(ws_b, "pmid")

    # Collect existing IDs and compute per-PMID counters
    pmid_counters = {}
    existing_ids = set()
    old_pattern = re.compile(r"^MB(\d{4})([A-Z]+)(\d{3})$")
    new_pattern = re.compile(r"^(\d+)_(\d{2})$")

    for r in range(2, ws_b.max_row + 1):
        v = ws_b.cell(row=r, column=b_id_col).value
        if v and str(v).strip():
            bid = str(v).strip()
            existing_ids.add(bid)
            m = new_pattern.match(bid)
            if m:
                p = m.group(1)
                idx = int(m.group(2))
                if p not in pmid_counters or idx > pmid_counters[p]:
                    pmid_counters[p] = idx
            # Also tolerate old MB* IDs for backward compat (ignore their counters)

    # Find rows needing IDs (iterate in order, assign per-PMID sequential indices)
    new_benchmark_ids = []
    for r in range(2, ws_b.max_row + 1):
        existing = ws_b.cell(row=r, column=b_id_col).value
        if existing and str(existing).strip():
            continue  # already has ID

        pmid_val = ws_b.cell(row=r, column=b_pmid_col).value
        pmid_str = str(int(pmid_val)) if pmid_val else "UNKNOWN"

        idx = pmid_counters.get(pmid_str, 0) + 1
        pmid_counters[pmid_str] = idx

        new_id = f"{pmid_str}_{idx:02d}"
        ws_b.cell(row=r, column=b_id_col, value=new_id)
        new_benchmark_ids.append(new_id)

    # ---- Case-level ----
    ws_c = wb["Case-level"]
    c_test_id_col = _find_header_col(ws_c, "test_id")
    c_bid_col = _find_header_col(ws_c, "benchmark_id")

    # Collect existing test_id counters per benchmark
    test_counters = {}
    for r in range(2, ws_c.max_row + 1):
        v = ws_c.cell(row=r, column=c_test_id_col).value
        if v and str(v).strip():
            tid = str(v).strip()
            # Extract benchmark_id prefix and test index
            # Pattern: <benchmark_id>_<NN> where benchmark_id may contain underscores
            m = re.match(r"^(.+)_(\d{2})$", tid)
            if m:
                bid_prefix = m.group(1)
                idx = int(m.group(2))
                if bid_prefix not in test_counters or idx > test_counters[bid_prefix]:
                    test_counters[bid_prefix] = idx

    new_test_count = 0
    for r in range(2, ws_c.max_row + 1):
        existing_test_id = ws_c.cell(row=r, column=c_test_id_col).value
        if existing_test_id and str(existing_test_id).strip():
            continue  # already has test_id

        existing_bid = ws_c.cell(row=r, column=c_bid_col).value
        if existing_bid and str(existing_bid).strip():
            bid = str(existing_bid).strip()
        else:
            continue  # No benchmark_id — needs manual mapping

        idx = test_counters.get(bid, 0) + 1
        test_counters[bid] = idx

        test_id = f"{bid}_{idx:02d}"
        ws_c.cell(row=r, column=c_test_id_col, value=test_id)
        new_test_count += 1

    wb.save(result_path)
    print(f"Assigned {len(new_benchmark_ids)} benchmark_ids: {new_benchmark_ids}")
    print(f"Assigned {new_test_count} test_ids")


# ---------------------------------------------------------------------------
# Propagate benchmark IDs to case rows (drug + cell_type matching)
# ---------------------------------------------------------------------------

def _tokenize_cell_type(text: str) -> set[str]:
    """Tokenize a cell type / cell context string into weighted tokens.

    Returns a set of lowercased tokens.  Tokens that look like cell-type
    identifiers (contain both letters and digits, e.g. "DAn1", "CD8+") are
    included as-is; pure alphabetic stopwords are excluded.
    """
    if not text:
        return set()
    # Split on common delimiters
    tokens = set()
    for tok in re.split(r"[\s,;/()]+", text.lower()):
        tok = tok.strip(".'\"-")
        if not tok:
            continue
        # Keep all non-trivial tokens
        if len(tok) <= 1:
            continue
        # Exclude very generic words that don't help disambiguate cell types
        if tok in ("the", "and", "from", "cell", "cells", "line", "lines",
                     "human", "mouse", "treated", "with", "for", "was", "were",
                     "using", "data", "sample", "samples", "group", "groups"):
            continue
        tokens.add(tok)
    return tokens


def _cell_type_match_score(bench_cell_context: str, case_cell_type: str) -> float:
    """Score how well a benchmark's cell_context matches a case's cell_type.

    Returns a float in [0, 1].  Tokens that look like specific cell-type
    identifiers (containing digits: "DAn1", "CD8+") carry 3x weight.
    """
    if not bench_cell_context or not case_cell_type:
        return 0.0

    b_toks = _tokenize_cell_type(bench_cell_context)
    c_toks = _tokenize_cell_type(case_cell_type)

    if not b_toks or not c_toks:
        return 0.0

    # Weighted: alphanumeric tokens (containing digits) are 3x
    def _is_specific(tok: str) -> bool:
        return bool(re.search(r"\d", tok))

    overlap_weight = 0
    total_weight = 0
    for tok in b_toks:
        w = 3 if _is_specific(tok) else 1
        total_weight += w
        if tok in c_toks:
            overlap_weight += w

    # Also count case tokens that aren't in benchmark tokens (for penalty)
    for tok in c_toks:
        if tok not in b_toks:
            w = 3 if _is_specific(tok) else 1
            total_weight += w

    if total_weight == 0:
        return 0.0
    return overlap_weight / total_weight


def _normalize_drug_name(name: str) -> str:
    """Normalize a drug name for comparison."""
    if not name:
        return ""
    return name.strip().lower()


def read_benchmark_ids(result_path: str) -> list[dict]:
    """Read all benchmark rows from result.xlsx and return them as dicts.

    Returns a list of dicts with keys: row, benchmark_id, pmid, dataset_accession,
    perturbation_name, cell_context. This is used by the two-phase write workflow:
    after assign-ids fills benchmark_ids, the LLM reads them back to know which
    benchmark_id goes with each case before writing case rows.
    """
    import openpyxl
    import os

    if not os.path.exists(result_path):
        print(f"Error: {result_path} not found", file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(result_path)
    ws_b = wb["Benchmark-level"]
    b_bid_col = _find_header_col(ws_b, "benchmark_id")
    b_pmid_col = _find_header_col(ws_b, "pmid")
    b_ds_col = _find_header_col(ws_b, "dataset_accession")
    b_pn_col = _find_header_col(ws_b, "perturbation_name")
    b_cc_col = _find_header_col(ws_b, "cell_context")

    rows = []
    for r in range(2, ws_b.max_row + 1):
        bid = ws_b.cell(row=r, column=b_bid_col).value
        pmid = ws_b.cell(row=r, column=b_pmid_col).value
        ds = ws_b.cell(row=r, column=b_ds_col).value
        drug = ws_b.cell(row=r, column=b_pn_col).value
        cc = ws_b.cell(row=r, column=b_cc_col).value
        if pmid or drug or cc:
            rows.append({
                "row": r,
                "benchmark_id": str(bid).strip() if bid else "",
                "pmid": str(pmid).strip() if pmid else "",
                "dataset_accession": str(ds).strip() if ds else "",
                "perturbation_name": str(drug).strip() if drug else "",
                "cell_context": str(cc).strip() if cc else "",
            })

    wb.close()
    return rows


def propagate_ids(result_path: str) -> None:
    """Propagate benchmark_id, pmid, dataset_accession from benchmark rows to
    case rows.

    Matches cases to benchmarks by drug (perturbation_name) + cell_type /
    cell_context, NOT by drug name alone.  When multiple benchmarks share the
    same drug, only the cell-type match disambiguates them.

    Prints a summary of assignments made, and WARNINGS for any ambiguous or
    low-confidence matches that need human review.
    """
    import openpyxl
    import os

    if not os.path.exists(result_path):
        print(f"Error: {result_path} not found", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(result_path)

    # ---- Read benchmark rows ----
    ws_b = wb["Benchmark-level"]
    b_bid_col = _find_header_col(ws_b, "benchmark_id")
    b_pmid_col = _find_header_col(ws_b, "pmid")
    b_ds_col = _find_header_col(ws_b, "dataset_accession")
    b_pn_col = _find_header_col(ws_b, "perturbation_name")
    b_cc_col = _find_header_col(ws_b, "cell_context")

    benchmarks = []  # list of {row, bid, pmid, dataset_accession, drug, cell_context}
    for r in range(2, ws_b.max_row + 1):
        bid = ws_b.cell(row=r, column=b_bid_col).value
        if not bid or not str(bid).strip():
            continue
        benchmarks.append({
            "row": r,
            "bid": str(bid).strip(),
            "pmid": str(ws_b.cell(row=r, column=b_pmid_col).value or "").strip(),
            "dataset_accession": str(ws_b.cell(row=r, column=b_ds_col).value or "").strip(),
            "drug": _normalize_drug_name(str(ws_b.cell(row=r, column=b_pn_col).value or "")),
            "cell_context": str(ws_b.cell(row=r, column=b_cc_col).value or "").strip(),
        })

    if not benchmarks:
        print("No benchmark rows with benchmark_id found — nothing to propagate.")
        wb.close()
        return

    # ---- Read case rows ----
    ws_c = wb["Case-level"]
    c_bid_col = _find_header_col(ws_c, "benchmark_id")
    c_pmid_col = _find_header_col(ws_c, "pmid")
    c_ds_col = _find_header_col(ws_c, "dataset_accession")
    c_drug_col = _find_header_col(ws_c, "drug")
    c_ct_col = _find_header_col(ws_c, "cell_type")

    assignments = []
    warnings = []
    for r in range(2, ws_c.max_row + 1):
        existing_bid = ws_c.cell(row=r, column=c_bid_col).value
        if existing_bid and str(existing_bid).strip():
            continue  # already has benchmark_id

        case_drug = _normalize_drug_name(str(ws_c.cell(row=r, column=c_drug_col).value or ""))
        case_cell_type = str(ws_c.cell(row=r, column=c_ct_col).value or "").strip()

        if not case_drug:
            warnings.append(f"Case row {r}: empty drug field — cannot propagate.")
            continue

        # Find candidate benchmarks: drug name must match EXACTLY
        candidates = [b for b in benchmarks if b["drug"] == case_drug]
        if not candidates:
            # Try substring match as fallback (one contains the other)
            candidates = [b for b in benchmarks
                          if case_drug in b["drug"] or b["drug"] in case_drug]

        if not candidates:
            warnings.append(
                f"Case row {r}: no benchmark found for drug '{case_drug}'. "
                f"Available drugs: {sorted(set(b['drug'] for b in benchmarks))}"
            )
            continue

        if len(candidates) == 1:
            best = candidates[0]
            reason = "only matching drug"
        else:
            # Multiple benchmarks with same drug — score by cell type match
            scored = [(b, _cell_type_match_score(b["cell_context"], case_cell_type))
                      for b in candidates]
            scored.sort(key=lambda x: x[1], reverse=True)
            best, best_score = scored[0]
            second_score = scored[1][1] if len(scored) > 1 else 0

            if best_score < 0.1:
                warnings.append(
                    f"Case row {r}: low cell-type match confidence "
                    f"(score={best_score:.2f}). "
                    f"Case cell_type='{case_cell_type}', "
                    f"best benchmark cell_context='{best['cell_context']}', "
                    f"drug='{case_drug}'. Manual review needed."
                )
            elif best_score < 0.3 and best_score - second_score < 0.15:
                warnings.append(
                    f"Case row {r}: ambiguous cell-type match "
                    f"(best_score={best_score:.2f}, second_score={second_score:.2f}). "
                    f"Case cell_type='{case_cell_type}', "
                    f"best benchmark='{best['bid']}' ({best['cell_context']}), "
                    f"runner-up='{scored[1][0]['bid']}' ({scored[1][0]['cell_context']}). "
                    f"Manual review needed."
                )
            reason = f"drug + cell_type match (score={best_score:.2f})"

        assignments.append((r, best, reason))

    # ---- Apply assignments ----
    for case_row, bench, reason in assignments:
        ws_c.cell(row=case_row, column=c_bid_col, value=bench["bid"])
        ws_c.cell(row=case_row, column=c_pmid_col, value=bench["pmid"])
        ws_c.cell(row=case_row, column=c_ds_col, value=bench["dataset_accession"])
        print(f"  Row {case_row}: → {bench['bid']} ({reason})")

    wb.save(result_path)

    print(f"\nPropagated {len(assignments)} case rows to benchmarks.")
    if warnings:
        print(f"\n⚠️  {len(warnings)} WARNING(S) — manual review needed:")
        for w in warnings:
            print(f"  - {w}")
    else:
        print("No warnings — all mappings look clean.")


# ---------------------------------------------------------------------------
# Validate case-to-benchmark mapping
# ---------------------------------------------------------------------------

def validate_mapping(result_path: str) -> int:
    """Validate that every case row is mapped to the correct benchmark.

    Checks performed:
      1. Every case row has a non-empty benchmark_id.
      2. The case's benchmark_id exists in the Benchmark-level sheet.
      3. The case's drug matches the assigned benchmark's perturbation_name.
      4. The case's cell_type is compatible with the benchmark's cell_context
         (not a clear mismatch).
      5. No other benchmark with the same drug has a substantially better
         cell_context match — catches cases mapped to the wrong benchmark
         when multiple benchmarks share a drug name.
      6. Every benchmark in Benchmark-level has at least one case in the
         Case-level sheet (no orphan benchmarks).

    Returns the number of errors found (0 = clean).
    """
    import openpyxl
    import os

    if not os.path.exists(result_path):
        print(f"Error: {result_path} not found", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(result_path)

    # ---- Read benchmark rows ----
    ws_b = wb["Benchmark-level"]
    b_bid_col = _find_header_col(ws_b, "benchmark_id")
    b_pn_col = _find_header_col(ws_b, "perturbation_name")
    b_cc_col = _find_header_col(ws_b, "cell_context")

    benchmarks = {}  # bid -> {drug, cell_context}
    for r in range(2, ws_b.max_row + 1):
        bid = ws_b.cell(row=r, column=b_bid_col).value
        if not bid or not str(bid).strip():
            continue
        bid = str(bid).strip()
        benchmarks[bid] = {
            "row": r,
            "drug": _normalize_drug_name(str(ws_b.cell(row=r, column=b_pn_col).value or "")),
            "cell_context": str(ws_b.cell(row=r, column=b_cc_col).value or "").strip(),
        }

    if not benchmarks:
        print("No benchmark rows with benchmark_id found — nothing to validate.")
        wb.close()
        return 0

    # ---- Read case rows ----
    ws_c = wb["Case-level"]
    c_bid_col = _find_header_col(ws_c, "benchmark_id")
    c_drug_col = _find_header_col(ws_c, "drug")
    c_ct_col = _find_header_col(ws_c, "cell_type")
    c_tg_col = _find_header_col(ws_c, "test_id")

    errors = []
    benchmarks_with_cases = set()  # track which benchmarks have ≥1 case
    for r in range(2, ws_c.max_row + 1):
        row_label = ws_c.cell(row=r, column=c_tg_col).value or f"row {r}"
        assigned_bid = ws_c.cell(row=r, column=c_bid_col).value
        case_drug = _normalize_drug_name(str(ws_c.cell(row=r, column=c_drug_col).value or ""))
        case_cell_type = str(ws_c.cell(row=r, column=c_ct_col).value or "").strip()

        # Check 1: benchmark_id must exist
        if not assigned_bid or not str(assigned_bid).strip():
            errors.append(f"[{row_label}] Missing benchmark_id — cannot validate mapping.")
            continue

        assigned_bid = str(assigned_bid).strip()

        # Check 2: assigned benchmark must exist in Benchmark-level sheet
        if assigned_bid not in benchmarks:
            errors.append(f"[{row_label}] benchmark_id '{assigned_bid}' not found in Benchmark-level sheet.")
            continue

        bench = benchmarks[assigned_bid]
        benchmarks_with_cases.add(assigned_bid)

        # Check 3: drug must match
        if case_drug and bench["drug"] and case_drug != bench["drug"]:
            # Try substring match
            if not (case_drug in bench["drug"] or bench["drug"] in case_drug):
                errors.append(
                    f"[{row_label}] Drug mismatch: case drug='{case_drug}' "
                    f"vs benchmark perturbation_name='{bench['drug']}' "
                    f"(benchmark_id={assigned_bid})."
                )
                continue

        # Check 4: cell_context must be compatible with cell_type
        if case_cell_type and bench["cell_context"]:
            score = _cell_type_match_score(bench["cell_context"], case_cell_type)
            if score < 0.05:
                errors.append(
                    f"[{row_label}] Cell-type mismatch: case cell_type='{case_cell_type}' "
                    f"vs benchmark cell_context='{bench['cell_context']}' "
                    f"(score={score:.2f}, benchmark_id={assigned_bid})."
                )

        # Check 5: no other benchmark with same drug is a substantially better match
        if case_drug and case_cell_type:
            same_drug_benchmarks = {
                bid: b for bid, b in benchmarks.items()
                if b["drug"] == case_drug or
                   (b["drug"] and case_drug and
                    (case_drug in b["drug"] or b["drug"] in case_drug))
            }
            if len(same_drug_benchmarks) > 1:
                scores = {
                    bid: _cell_type_match_score(b["cell_context"], case_cell_type)
                    for bid, b in same_drug_benchmarks.items()
                }
                best_bid = max(scores, key=scores.get)
                assigned_score = scores.get(assigned_bid, 0)
                best_score = scores[best_bid]

                # Flag if a DIFFERENT benchmark for the same drug has a
                # substantially better cell-type match (≥0.15 absolute gap
                # AND the assigned score is low)
                if (best_bid != assigned_bid
                        and best_score - assigned_score >= 0.15
                        and assigned_score < 0.4):
                    errors.append(
                        f"[{row_label}] Likely wrong benchmark mapping: "
                        f"case drug='{case_drug}', cell_type='{case_cell_type}'. "
                        f"Assigned to {assigned_bid} "
                        f"(cell_context='{benchmarks[assigned_bid]['cell_context']}', "
                        f"score={assigned_score:.2f}) "
                        f"but {best_bid} "
                        f"(cell_context='{benchmarks[best_bid]['cell_context']}', "
                        f"score={best_score:.2f}) is a much better match."
                    )

    # Check 6: every benchmark must have at least one case (no orphan benchmarks)
    for bid in benchmarks:
        if bid not in benchmarks_with_cases:
            b = benchmarks[bid]
            errors.append(
                f"[Benchmark-level row {b['row']}] Orphan benchmark: {bid} "
                f"(perturbation_name='{b['drug']}', cell_context='{b['cell_context']}') "
                f"has zero cases in Case-level sheet."
            )

    wb.close()

    if errors:
        print(f"\n❌ VALIDATION FAILED — {len(errors)} issue(s) found:\n")
        for e in errors:
            print(f"  {e}")
        print(f"\n  Fix these issues before proceeding.")
        return len(errors)
    else:
        print("OK — all case-to-benchmark mappings validated successfully.")
        return 0


# ---------------------------------------------------------------------------
# Caption extraction
# ---------------------------------------------------------------------------

# Patterns that mark the start of a figure or table caption line.
# Examples: "Figure 1.", "Fig. 1.", "Figure S1", "Figure A1", "Figure 3:",
# "# Figure 3.", "Table S8.", "Table A1.", "Supplemental Figure 1",
# "Supplemental Fig. 2", "Extended Data Fig. 4f"
#
# [A-Z]?\d+ covers: no prefix (Figure 1), supplementary (Figure S1),
# and appendix (Figure A1, Table A1).
# [A-Z]?(?![a-zA-Z]) after \d+ matches sub-figure labels like "Figure 1A"
# but prevents eating the first letter of caption words.
# [.:]? accepts both "Figure 1." and "Figure 1:" separators.
_CAPTION_MARKER = re.compile(
    r"^(?:#+\s*)?"
    r"(?:(?:Figure|Fig\.)\s+[A-Z]?\d+[A-Z]?(?![a-zA-Z])[.:]?(?:\s|$)|"
    r"(?:Supplemental\s+(?:Figure|Fig\.))\s+\d+[A-Z]?(?![a-zA-Z])[.:]?(?:\s|$)|"
    r"Extended\s+Data\s+Fig\.?\s+\d+[A-Za-z][.:]?(?:\s|$)|"
    r"Table\s+[A-Z]?\d+[A-Z]?(?![a-zA-Z])[.:]?(?:\s|$))",
    re.IGNORECASE,
)

# Lines that signal the end of a caption block.
# Image lines (![...]) are NOT breaks — they are skipped during collection
# because figures commonly interleave images between title and panel descriptions.
_CAPTION_BREAK = re.compile(
    r"^####|^<table>|^<\/table>",
)

# Caption panel markers: "(A)", "(B and C)", "(1)", "(i)", etc.
_CAPTION_PANEL = re.compile(r"^\([A-Za-z0-9]")


def extract_captions(input_md: str, output_md: str) -> None:
    """Extract all figure/table captions from a paper markdown and write to a file.

    Each caption is written as a markdown block prefixed with its marker line
    for easy reference (e.g. ``### Figure 1. ...``).
    """
    import os

    with open(input_md, "r", encoding="utf-8") as f:
        lines = f.readlines()

    captions = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        m = _CAPTION_MARKER.match(line)
        if m:
            marker = line
            caption_lines = []
            # Include text on the marker line itself (after the figure/table number).
            # Single-line captions (e.g. Supplemental Figures) carry all their text
            # on the marker line; only collecting subsequent lines would miss them.
            tail = line[m.end():].strip()
            if tail:
                caption_lines.append(tail)
            i += 1
            # Collect caption text
            while i < len(lines):
                nxt = lines[i].strip()

                # Stop at next caption marker
                if _CAPTION_MARKER.match(nxt):
                    break
                # Stop at explicit breaks (section headers within body, tables)
                if _CAPTION_BREAK.match(nxt):
                    break
                # Skip image lines — they are part of the figure, not breaks
                if nxt.startswith("!["):
                    i += 1
                    continue
                # Stop at blank line that precedes body text or a section header.
                # Blank lines between panel descriptions are normal; only break
                # when the following non-blank line looks like body text (not a
                # caption panel marker like "(A)" or an image).
                if not nxt:
                    # Look ahead to next non-blank line
                    j = i + 1
                    while j < len(lines) and not lines[j].strip():
                        j += 1
                    if j < len(lines):
                        ahead = lines[j].strip()
                        # Section header → end of caption
                        if ahead.startswith("#"):
                            break
                        # Next caption marker → end of this caption
                        if _CAPTION_MARKER.match(ahead):
                            break
                        # Image → still part of the figure, skip this blank
                        if ahead.startswith("!["):
                            i += 1
                            continue
                        # Line does NOT look like a caption panel → body text
                        if not _CAPTION_PANEL.match(ahead):
                            break
                    i += 1
                    continue

                if nxt:
                    caption_lines.append(nxt)
                i += 1

            caption_text = " ".join(caption_lines)
            if caption_text:
                # Clean up the marker for heading display (strip leading #)
                clean_marker = marker.lstrip("#").strip()
                captions.append((clean_marker, caption_text))
        else:
            i += 1

    # Write output
    os.makedirs(os.path.dirname(output_md) or ".", exist_ok=True)
    with open(output_md, "w", encoding="utf-8") as f:
        f.write("# Figure and Table Captions\n\n")
        if not captions:
            f.write("*(no captions found)*\n")
        for marker, text in captions:
            f.write(f"### {marker}\n\n{text}\n\n")

    print(f"Extracted {len(captions)} captions to {output_md}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Benchmark extraction helpers")
    sub = parser.add_subparsers(dest="command")

    # smiles
    p_smiles = sub.add_parser("smiles", help="Look up SMILES for a drug name")
    p_smiles.add_argument("drug_name")

    # pubmed
    p_pubmed = sub.add_parser("pubmed", help="Fetch PubMed metadata by PMID")
    p_pubmed.add_argument("pmid")

    # geo
    p_geo = sub.add_parser("geo", help="Fetch GEO series metadata")
    p_geo.add_argument("accession")

    # gen-benchmark-id
    p_bid = sub.add_parser("gen-benchmark-id", help="Generate next benchmark_id")
    p_bid.add_argument("pmid")
    p_bid.add_argument("result_path")

    # gen-test-id
    p_tid = sub.add_parser("gen-test-id", help="Generate next test_id")
    p_tid.add_argument("benchmark_id")
    p_tid.add_argument("result_path")

    # write-xlsx
    p_write = sub.add_parser("write-xlsx", help="Append benchmark/case rows to result.xlsx (auto-creates if missing)")
    p_write.add_argument("result_path")
    p_write.add_argument("json_data", help='JSON string: {"benchmarks": [...], "cases": [...]}')

    # assign-ids
    p_assign = sub.add_parser("assign-ids", help="Assign benchmark_ids and test_ids to rows that lack them")
    p_assign.add_argument("result_path")

    # extract-captions
    p_cap = sub.add_parser("extract-captions", help="Extract all figure/table captions from paper markdown")
    p_cap.add_argument("input_md", help="Path to paper markdown file")
    p_cap.add_argument("output_md", help="Path to output captions markdown file")

    # read-benchmark-ids
    p_read = sub.add_parser("read-benchmark-ids", help="Read all benchmark rows from result.xlsx (used to resolve benchmark_ids after assign-ids)")
    p_read.add_argument("result_path")

    # propagate-ids
    p_prop = sub.add_parser("propagate-ids", help="Propagate benchmark_id to case rows using drug + cell_type matching (recovery tool)")
    p_prop.add_argument("result_path")

    # validate-mapping
    p_val = sub.add_parser("validate-mapping", help="Validate case-to-benchmark mapping; reports wrong/ambiguous assignments")
    p_val.add_argument("result_path")

    args = parser.parse_args()

    if args.command == "smiles":
        result = fetch_smiles(args.drug_name)
        print(result if result else "(not found)")

    elif args.command == "pubmed":
        result = fetch_pubmed_metadata(args.pmid)
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.command == "geo":
        result = fetch_geo_metadata(args.accession)
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.command == "gen-benchmark-id":
        result = generate_benchmark_id(args.pmid, args.result_path)
        print(result)

    elif args.command == "gen-test-id":
        result = generate_test_id(args.benchmark_id, args.result_path)
        print(result)

    elif args.command == "write-xlsx":
        data = json.loads(args.json_data)
        write_to_xlsx(args.result_path, data.get("benchmarks", []), data.get("cases", []))

    elif args.command == "assign-ids":
        assign_ids(args.result_path)

    elif args.command == "extract-captions":
        extract_captions(args.input_md, args.output_md)

    elif args.command == "read-benchmark-ids":
        rows = read_benchmark_ids(args.result_path)
        print(json.dumps(rows, ensure_ascii=False, indent=2))

    elif args.command == "propagate-ids":
        propagate_ids(args.result_path)

    elif args.command == "validate-mapping":
        err_count = validate_mapping(args.result_path)
        if err_count > 0:
            sys.exit(1)

    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
